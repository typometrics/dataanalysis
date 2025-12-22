"""
Verb-Centered Constituent Size Analysis.

This module provides comprehensive analysis of constituent sizes around verbs,
with support for horizontal and diagonal growth factors.

REFACTORED: Now uses modular architecture with separate components for
computation, layout, and formatting.
"""

import numpy as np
import os
from typing import Dict, Optional, Union

# Import refactored components
try:
    from verb_centered_model import TableConfig, TableStructure, GridCell
    from verb_centered_builder import VerbCenteredTableBuilder
    from verb_centered_formatters import (
        TextTableFormatter, TSVFormatter, ExcelFormatter, convert_table_to_grid_cells
    )
    REFACTORED_MODULES_AVAILABLE = True
except ImportError as e:
    print(f"Warning: Refactored modules not fully available: {e}")
    REFACTORED_MODULES_AVAILABLE = False
    # Define minimal GridCell for backward compatibility
    class GridCell:
        def __init__(self, text="", value=None, cell_type='normal', rich_text=None, is_astonishing=False):
            self.text = text
            self.value = value
            self.cell_type = cell_type
            self.rich_text = rich_text if rich_text is not None else []
            self.is_astonishing = is_astonishing


# ============================================================================
# NEW PUBLIC API (Refactored)
# ============================================================================

def create_verb_centered_table(
    position_averages: Dict[str, float],
    config: Optional[TableConfig] = None,
    ordering_stats: Optional[Dict] = None,
    **kwargs
) -> Optional[TableStructure]:
    """
    Create table structure (unified entry point).
    
    This is the new, preferred API for creating verb-centered tables.
    
    Args:
        position_averages: Dictionary of average sizes and factors
        config: Table configuration (uses defaults if None)
        ordering_stats: Optional ordering statistics
        **kwargs: Backward compatibility - can pass individual config params
        
    Returns:
        TableStructure object that can be formatted multiple ways, or None if modules unavailable
        
    Example:
        >>> config = TableConfig(show_horizontal_factors=True, arrow_direction='left_to_right')
        >>> table = create_verb_centered_table(position_averages, config, ordering_stats)
        >>> text_output = TextTableFormatter(table).format()
        >>> ExcelFormatter().save(table, 'output.xlsx')
    """
    if not REFACTORED_MODULES_AVAILABLE:
        print("Warning: Using legacy implementation")
        return None
    
    # Build config from kwargs if not provided
    if config is None:
        config = TableConfig(
            show_horizontal_factors=kwargs.get('show_horizontal_factors', False),
            show_diagonal_factors=kwargs.get('show_diagonal_factors', False),
            show_ordering_triples=kwargs.get('show_ordering_triples', False),
            show_row_averages=kwargs.get('show_row_averages', False),
            show_marginal_means=kwargs.get('show_marginal_means', True),
            arrow_direction=kwargs.get('arrow_direction', 'diverging')
        )
    
    builder = VerbCenteredTableBuilder(position_averages, config, ordering_stats)
    return builder.build()


# ============================================================================
# LEGACY PUBLIC API (Backward Compatible)
# ============================================================================

def compute_average_sizes_table(all_langs_average_sizes_filtered):
    """
    Compute average constituent sizes at each position for different totals.
    Also computes GEOMETRIC MEANS of ratios between positions (Growth Factors).
    
    Returns a dictionary of averages keyed by position string (e.g. 'right_1_totright_2').
    Growth factors are keyed as 'factor_{key_B}_vs_{key_A}'.
    """
    # Dictionary to store sums and counts for averaging sizes (Arithmetic Mean)
    position_sums = {}
    position_counts = {}
    
    # Dictionary to store sums of logs and counts for averaging ratios (Geometric Mean)
    # key: (key_B, key_A) -> {'sum_log_ratio': 0.0, 'count': 0}
    ratio_stats = {}
    
    # Define pairs of positions to compare for Growth Factors
    # 1. Horizontal Right (Pos N vs Pos N-1)
    # 2. Horizontal Left (Pos N vs Pos N-1) -> Note: Logic is usually Outer/Inner comparisons
    # 3. XVX
    # 4. Diagonals (Right Tot vs Tot-1)
    # 5. Diagonals (Left Tot vs Tot+1)
    
    pairs_to_track = []
    
    # Horizontal Right: right_{pos}_totright_{tot} vs right_{pos-1}_totright_{tot}
    for tot in range(1, 5):
        for pos in range(2, tot + 1):
            key_b = f'right_{pos}_totright_{tot}'
            key_a = f'right_{pos-1}_totright_{tot}'
            pairs_to_track.append((key_b, key_a, 'diverging')) # val / prev_val

    # Horizontal Left: left_{pos}_totleft_{tot} vs left_{pos-1}_totleft_{tot}
    # In extract_grid/format logic:
    # If arrow rightwards: prev_val / val (Size(pos-1)/Size(pos))
    # If diverging: val / prev_val (Size(pos)/Size(pos-1))
    # Let's track BOTH simply as B/A for now and apply direction logic later?
    # No, we need to know WHICH ratio to average.
    # Standard usually is "Growth" = Outer/Inner or Distal/Proximal.
    # Diverging: V -> R1 -> R2. Ratio R2/R1.
    # Diverging: L2 <- L1 <- V. Ratio L2/L1 (Outer/Inner).
    # key_b=Left_Pos(outer), key_a=Left_Pos-1(inner).
    # Matches 'val/prev_val' logic where 'val' is current loop pos (outer).
    for tot in range(1, 5):
        for pos in range(2, tot + 1):
            key_b = f'left_{pos}_totleft_{tot}'
            key_a = f'left_{pos-1}_totleft_{tot}'
            pairs_to_track.append((key_b, key_a, 'diverging'))

    # XVX: right_1 vs left_1
    pairs_to_track.append(('xvx_right_1', 'xvx_left_1', 'diverging'))

    # Diagonals Right: right_{pos+1}_totright_{tot} vs right_{pos}_totright_{tot-1}
    for tot in range(2, 5):
        for pos in range(1, tot): # pos in source tot-1 (1..tot-1)
            key_a = f'right_{pos}_totright_{tot-1}' # source
            key_b = f'right_{pos+1}_totright_{tot}' # target
            pairs_to_track.append((key_b, key_a, 'diverging'))

    # Diagonals Left: left_{pos}_totleft_{tot} vs left_{pos+1}_totleft_{tot+1}
    # Wait, extraction logic: tgt / src.
    # tgt = left_{pos}_totleft_{tot}
    # src = left_{pos+1}_totleft_{tot+1}
    # Ratio = tgt / src.
    for tot in range(1, 4):
         for pos in range(1, tot + 1):
            key_b = f'left_{pos}_totleft_{tot}' # target
            key_a = f'left_{pos+1}_totleft_{tot+1}' # source
            pairs_to_track.append((key_b, key_a, 'diverging'))
    
    
    # Collect all values across languages
    for lang, positions in all_langs_average_sizes_filtered.items():
        # 1. Accumulate sizes
        for position_key, value in positions.items():
            if position_key not in position_sums:
                position_sums[position_key] = 0
                position_counts[position_key] = 0
            if value > 0:
                position_sums[position_key] += np.log(value)
            else:
                pass # handle 0 size? Should not happen with GM.
            position_counts[position_key] += 1
            
        # 2. Accumulate ratios
        for key_b, key_a, direction in pairs_to_track:
            val_b = positions.get(key_b)
            val_a = positions.get(key_a)
            
            if val_b is not None and val_a is not None and val_a > 0 and val_b > 0:
                # Compute ratio
                ratio = val_b / val_a
                
                pair_key = (key_b, key_a)
                if pair_key not in ratio_stats:
                    ratio_stats[pair_key] = {'sum_log_ratio': 0.0, 'count': 0}
                
                ratio_stats[pair_key]['sum_log_ratio'] += np.log(ratio)
                ratio_stats[pair_key]['count'] += 1

    # Calculate averages
    results = {}
    
    # Geometric Means for Sizes (Cross-Language)
    for position_key in position_sums:
        if position_counts[position_key] > 0:
            results[position_key] = np.exp(position_sums[position_key] / position_counts[position_key])
            
    # Geometric Means for Ratios
    for pair_key, stats in ratio_stats.items():
        if stats['count'] > 0:
            avg_log = stats['sum_log_ratio'] / stats['count']
            geo_mean_ratio = np.exp(avg_log)
            
            # Store with a clean key
            key_b, key_a = pair_key
            factor_key = f'factor_{key_b}_vs_{key_a}'
            results[factor_key] = geo_mean_ratio
            
    return results


# ============================================================================
# SIMPLIFIED WRAPPER FUNCTIONS (Using Refactored Components)
# ============================================================================

def format_verb_centered_table(
    position_averages: Dict[str, float],
    show_horizontal_factors: bool = False,
    show_diagonal_factors: bool = False,
    show_ordering_triples: bool = False,
    show_row_averages: bool = False,
    show_marginal_means: bool = True,
    arrow_direction: str = 'diverging',
    ordering_stats: Optional[Dict] = None,
    disorder_percentages: Optional[Dict] = None,
    save_tsv: bool = False,
    output_dir: str = 'data',
    filename: Optional[str] = None
) -> str:
    """
    Format verb-centered table as text (legacy API).
    
    Now routes through refactored components for cleaner implementation.
    Falls back to old implementation if refactored modules unavailable.
    
    Args:
        position_averages: Dictionary of average sizes and factors
        show_horizontal_factors: Show horizontal growth factors
        show_diagonal_factors: Show diagonal growth factors
        show_ordering_triples: Show ordering statistics
        show_row_averages: Show row averages (GM, N, Slope)
        show_marginal_means: Show marginal mean rows
        arrow_direction: Arrow direction mode ('diverging', 'inward', 'left_to_right', 'right_to_left')
        ordering_stats: Optional ordering statistics
        disorder_percentages: (Deprecated) Use ordering_stats instead
        save_tsv: Whether to save TSV file
        output_dir: Directory for TSV output
        filename: Filename for TSV output
        
    Returns:
        Formatted text string
    """
    if REFACTORED_MODULES_AVAILABLE:
        # Use new implementation
        config = TableConfig(
            show_horizontal_factors=show_horizontal_factors,
            show_diagonal_factors=show_diagonal_factors,
            show_ordering_triples=show_ordering_triples,
            show_row_averages=show_row_averages,
            show_marginal_means=show_marginal_means,
            arrow_direction=arrow_direction
        )
        
        table = create_verb_centered_table(position_averages, config, ordering_stats)
        if table is None:
            return "Error: Could not create table"
        
        # Format as text
        text_output = TextTableFormatter(table).format()
        
        # Save TSV if requested
        if save_tsv:
            if filename is None:
                parts = ['verb_centered_table']
                if show_horizontal_factors:
                    parts.append('factors')
                if show_diagonal_factors:
                    parts.append('diagonal')
                if show_horizontal_factors or show_diagonal_factors:
                    parts.append(arrow_direction)
                filename = "_".join(parts) + ".tsv"
            
            tsv_path = os.path.join(output_dir, filename)
            TSVFormatter().save(table, tsv_path)
            text_output += f"\n\nTable saved to: {tsv_path}"
        
        return text_output
    
    else:
        # Fall back to original implementation
        return _format_verb_centered_table_legacy(
            position_averages, show_horizontal_factors, show_diagonal_factors,
            arrow_direction, ordering_stats, show_ordering_triples,
            show_row_averages, show_marginal_means, disorder_percentages,
            save_tsv, output_dir, filename
        )


def extract_verb_centered_grid(
    position_averages: Dict[str, float],
    show_horizontal_factors: bool = False,
    show_diagonal_factors: bool = False,
    arrow_direction: str = 'diverging',
    ordering_stats: Optional[Dict] = None,
    show_ordering_triples: bool = False,
    show_row_averages: bool = False,
    show_marginal_means: bool = True
):
    """
    Constructs a grid of GridCell objects representing the table (legacy API).
    
    Now routes through refactored components.
    Falls back to old implementation if refactored modules unavailable.
    
    Args:
        position_averages: Dictionary of average sizes and factors
        show_horizontal_factors: Show horizontal growth factors
        show_diagonal_factors: Show diagonal growth factors
        arrow_direction: Arrow direction mode
        ordering_stats: Optional ordering statistics
        show_ordering_triples: Show ordering statistics
        show_row_averages: Show row averages
        show_marginal_means: Show marginal mean rows
        
    Returns:
        List of rows, each row is a list of GridCell objects
    """
    if REFACTORED_MODULES_AVAILABLE:
        # Use new implementation
        config = TableConfig(
            show_horizontal_factors=show_horizontal_factors,
            show_diagonal_factors=show_diagonal_factors,
            show_ordering_triples=show_ordering_triples,
            show_row_averages=show_row_averages,
            show_marginal_means=show_marginal_means,
            arrow_direction=arrow_direction
        )
        
        table = create_verb_centered_table(position_averages, config, ordering_stats)
        if table is None:
            return []
        
        # Convert to old GridCell format
        return convert_table_to_grid_cells(table)
    
    else:
        # Fall back to original implementation
        return _extract_verb_centered_grid_legacy(
            position_averages, show_horizontal_factors, show_diagonal_factors,
            arrow_direction, ordering_stats, show_ordering_triples,
            show_row_averages, show_marginal_means
        )


def save_excel_verb_centered_table(grid_rows, output_path):
    """
    Save verb-centered table to Excel file (legacy API).
    
    Now routes through refactored components when possible.
    
    Args:
        grid_rows: List of GridCell rows (legacy format) or TableStructure
        output_path: Path to output .xlsx file
    """
    if REFACTORED_MODULES_AVAILABLE:
        # Check if grid_rows is already a TableStructure
        if isinstance(grid_rows, TableStructure):
            ExcelFormatter().save(grid_rows, output_path)
        else:
            # It's a list of GridCell rows - use old implementation
            _save_excel_verb_centered_table_legacy(grid_rows, output_path)
    else:
        # Fall back to original implementation
        _save_excel_verb_centered_table_legacy(grid_rows, output_path)


# ============================================================================
# LEGACY IMPLEMENTATION (Preserved for Fallback)
# ============================================================================

def get_growth_direction(side, arrow_direction='diverging'):
    """
    Determines if the computation should be 'outward' (Center->Periphery) or 'inward'.
    """
    if arrow_direction in ['diverging', 'outward']:
        return 'outward'
    elif arrow_direction in ['inward', 'converging']:
        return 'inward'
    elif arrow_direction == 'left_to_right':
        # Left side: L4->L1 is flowing Right (towards center). Inward.
        # Right side: R1->R4 is flowing Right (away from center). Outward.
        return 'inward' if side == 'left' else 'outward'
    elif arrow_direction == 'right_to_left':
        # Left side: L1->L4 is flowing Left (away from center). Outward.
        # Right side: R4->R1 is flowing Left (towards center). Inward.
        return 'outward' if side == 'left' else 'inward'
    return 'outward'

def get_arrow(side, arrow_type, arrow_direction='diverging'):
    """
    Returns the appropriate arrow for a factor.
    """
    if arrow_direction == 'left_to_right':
        return '→' if arrow_type == 'horizontal' else '↗'
    elif arrow_direction == 'right_to_left':
        return '←' if arrow_type == 'horizontal' else '↙'
    
    # Legacy modes
    is_outward = get_growth_direction(side, arrow_direction) == 'outward'
    
    if side == 'right':
        if arrow_type == 'horizontal':
            return '→' if is_outward else '←'
        else: # diagonal
            return '↗' if is_outward else '↙'
    else: # left
        if arrow_type == 'horizontal':
            return '←' if is_outward else '→'
        else: # diagonal
            return '↙' if is_outward else '↗'

def _extract_verb_centered_grid_legacy(position_averages, 
                               show_horizontal_factors=False, 
                               show_diagonal_factors=False,
                               arrow_direction='diverging',
                               ordering_stats=None,
                               show_ordering_triples=False,
                               show_row_averages=False,
                               show_marginal_means=True):
    """
    LEGACY IMPLEMENTATION - Constructs a grid of GridCell objects representing the table.
    """
    rows = []
    is_outward = arrow_direction in ['diverging', 'rightwards', 'outward']
    
    # Generate Header Row
    header_row = [GridCell("Row", cell_type='label')]
    header_row.append(GridCell("")) # Mirrored Extra Column
    
    # Left Headers (Fixed 7 slots in data logic)
    l_headers = [GridCell("") for _ in range(7)]
    if show_horizontal_factors:
         # Logic matches loop: 6 - (pos-1)*2.
         # Pos 4 (L4): 0. Pos 3 (L3): 2. Pos 2 (L2): 4. Pos 1 (L1): 6.
         for pos in [4, 3, 2, 1]:
             idx = 6 - (pos - 1) * 2
             if 0 <= idx < 7:
                 l_headers[idx] = GridCell(f"L{pos}", cell_type='label')
    else:
         # Logic matches loop: 3 - (pos - 1).
         # Pos 4: 0. Pos 1: 3.
         for pos in [4, 3, 2, 1]:
             idx = 3 - (pos - 1)
             if 0 <= idx < 7:
                 l_headers[idx] = GridCell(f"L{pos}", cell_type='label')
                 
    header_row.extend(l_headers)
    header_row.append(GridCell("V", cell_type='label'))
    
    # Right Headers (Max Tot 4)
    for pos in range(1, 5):
        header_row.append(GridCell(f"R{pos}", cell_type='label'))
        if pos < 4 and show_horizontal_factors:
            header_row.append(GridCell("", cell_type='factor'))
            
    if show_row_averages:
        header_row.append(GridCell("[GM | N | Slope]", cell_type='comment'))
    
    rows.append(header_row)
    
    # Constants
    # E_VAL = GridCell(" " * 11) 
    
    # Colors
    COLOR_RED = "FF0000"
    COLOR_GREY = "555555"
    COLOR_DARK = "000000"
    COLOR_BLUE = "0000FF"
    COLOR_ORANGE = "FFA500"

    # Helper for Marginal GM
    def calc_marginal_gm(keys):
        vals = []
        for k in keys:
            v = position_averages.get(k)
            if v is not None and v > 0:
                vals.append(v)
        if not vals: return None
        return np.exp(np.mean(np.log(vals)))

    # ---------------------------------------------------------
    # 0. Right Marginal Means (Top)
    # ---------------------------------------------------------
    if show_marginal_means:
        calc_gm = calc_marginal_gm # alias
        
        # -----------------------------------------------
        # 0A. Vertical Means (Row 1)
        # -----------------------------------------------
        # Col Ri: GM(Sizes in Column Ri).
        # Col Fij: GM(Horizontal Factors in Col Fij) AND GM(Diagonal Factors arriving at Col Ri+1?).
        #   - H-factors at Fij connect Ri -> R(i+1). 
        #   - D-factors displayed in grid at Fij connect Ri(T-1) -> R(i+1)(T).
        #   So these naturally align in the "Fij" column.
        
        v_size_means = {}
        v_h_fac_means = {}
        v_d_fac_means = {}
        
        for pos in range(1, 5):
            # Size Mean at Pos
            keys = [f'right_{pos}_totright_{tot}' for tot in range(pos, 5)]
            v_size_means[pos] = calc_gm(keys)
            
            # Factor Means at Pos (to right of pos)
            # Factor connects Pos -> Pos+1
            if pos < 4:
                # H-Factor keys
                # factor_right_{pos+1}_tot_{tot}_vs_right_{pos}_tot_{tot}
                h_keys = []
                # D-Factor keys
                # factor_right_{pos+1}_tot_{tot}_vs_right_{pos}_tot_{tot-1}
                d_keys = []
                
                for tot in range(pos + 1, 5): # Need tot for target
                    # Horizontal: tot range (pos+1..4)
                    k_b = f'right_{pos+1}_totright_{tot}'
                    k_a = f'right_{pos}_totright_{tot}'
                    h_keys.append(f'factor_{k_b}_vs_{k_a}')
                    
                    # Diagonal: tot range (pos+1..4)?
                    # D-factor connects src (tot-1) to tgt (tot).
                    # Valid if src exists. src is pos at tot-1.
                    # tgt is pos+1 at tot.
                    k_a_diag = f'right_{pos}_totright_{tot-1}'
                    d_keys.append(f'factor_{k_b}_vs_{k_a_diag}')
                    
                v_h_fac_means[pos] = calc_gm(h_keys)
                v_d_fac_means[pos] = calc_gm(d_keys)

        # Construct Row 1: [M Vert]
        # Label
        v_row = [GridCell("M Vert", cell_type='label', rich_text=[("M Vert", COLOR_BLUE, True)])]
        v_row.append(GridCell("")) # Mirror Extra Empty
        for _ in range(7): v_row.append(GridCell(""))
        v_row.append(GridCell("")) # V (Empty)
        
        for pos in range(1, 5):
            # Value Cell
            gm = v_size_means.get(pos)
            if gm is not None:
                txt = f"{gm:.3f}"
                v_row.append(GridCell(txt, value=float(gm), cell_type='value', rich_text=[(txt, COLOR_BLUE, True)]))
            else:
                v_row.append(GridCell(""))
                
            # Factor Cell
            if pos < 4 and show_horizontal_factors: # Only up to R3->R4
                h_gm = v_h_fac_means.get(pos)
                d_gm = v_d_fac_means.get(pos)
                
                parts = []
                rich = []
                
                if h_gm is not None:
                    val = h_gm if is_outward else (1.0 / h_gm if h_gm != 0 else 1.0)
                    sym = get_arrow('right', 'horizontal', arrow_direction)
                    s = f"×{val:.2f}{sym}"
                    parts.append(s)
                    rich.append((s, COLOR_BLUE, False))
                
                # D-Factor removed for Vertical Row (H-only)
                    
                if parts:
                    v_row.append(GridCell(" ".join(parts), cell_type='factor', rich_text=rich))
                else:
                    v_row.append(GridCell(""))
        
        v_row.append(GridCell("")) # Extra Col for M Vert
        rows.append(v_row)

        # -----------------------------------------------
        # 0B. Diagonal Means (Row 2)
        # -----------------------------------------------
        # Map Diagonals to Columns.
        # R4 (Outer) -> Diag 0 (Last X).
        # R1 (Inner) -> Diag 3 (First X).
        # Mapping: Diag_Index_for_Pos(P) = 4 - P.  (P=4->0, P=1->3)
        
        d_size_means = {}
        d_h_fac_means = {}
        d_d_fac_means = {}
        
        # Pre-calculate means for all Diagonals (k=0..3)
        # k=0: Last X. k=3: First X.
        diag_size_raw = {}
        diag_h_raw = {} 
        diag_d_raw = {}
        

        # 0B. Diagonal Means (Row 2) - New Logic
        # -----------------------------------------------
        d_row = [GridCell("M Diag", cell_type='label', rich_text=[("M Diag", COLOR_ORANGE, True)])]
        d_row.append(GridCell("")) # Mirror Extra Empty
        for _ in range(7): d_row.append(GridCell("")) # Left
        d_row.append(GridCell("")) # V (Empty)
        
        # Calculate Specific Diagonal Means
        # 1. Longest Diagonal (Len 3): R1T1..R4T4 (Diag 0)
        # u=2..4. t=u.
        keys_fac_d3 = []
        keys_sz_d3 = [f'right_1_totright_1', f'right_2_totright_2', f'right_3_totright_3', f'right_4_totright_4']
        for u in range(2, 5):
             keys_fac_d3.append(f'factor_right_{u}_totright_{u}_vs_right_{u-1}_totright_{u-1}')
        gm_fac_d3 = calc_gm(keys_fac_d3)
        gm_sz_d3 = calc_gm(keys_sz_d3)
        
        # 2. Medium Diagonal (Len 2): R1T2..R3T4 (Diag 1)
        # u=2..3. t=u+1.
        keys_fac_d2 = []
        keys_sz_d2 = [f'right_1_totright_2', f'right_2_totright_3', f'right_3_totright_4']
        for u in range(2, 4):
             keys_fac_d2.append(f'factor_right_{u}_totright_{u+1}_vs_right_{u-1}_totright_{u}')
        gm_fac_d2 = calc_gm(keys_fac_d2)
        gm_sz_d2 = calc_gm(keys_sz_d2)
        
        # 3. Shortest Diagonal (Len 1): R1T3..R2T4 (Diag 2)
        # u=2. t=u+2.
        keys_fac_d1 = []
        keys_sz_d1 = [f'right_1_totright_3', f'right_2_totright_4']
        keys_fac_d1.append(f'factor_right_{2}_totright_{4}_vs_right_{1}_totright_{3}')
        gm_fac_d1 = calc_gm(keys_fac_d1)
        gm_sz_d1 = calc_gm(keys_sz_d1)
        
        # Fill Right Grid (8 Cols: 7 + Extra)
        # 0: R1 Val (Empty)
        # 1: F12 (Empty)
        # 2: R2 Val (Empty)
        # 3: F23 (Mean Diag Len 1)
        # 4: R3 Val (Empty)
        # 5: F34 (Mean Diag Len 2)
        # 6: R4 Val (Empty)
        # 7: Extra (Mean Diag Len 3)
        
        right_cells = [GridCell("") for _ in range(8)]
        
        def make_fac(fac, sz):
            if fac is None: return GridCell("")
            val = fac if is_outward else (1.0 / fac if fac != 0 else 1.0)
            sym = get_arrow('right', 'diagonal', arrow_direction)
            s = f"{sz:.2f} ×{val:.2f}{sym}" if sz is not None else f"×{val:.2f}{sym}"
            return GridCell(s, cell_type='factor', rich_text=[(s, COLOR_ORANGE, False)])
            
        right_cells[3] = make_fac(gm_fac_d1, gm_sz_d1)
        right_cells[5] = make_fac(gm_fac_d2, gm_sz_d2)
        right_cells[7] = make_fac(gm_fac_d3, gm_sz_d3)
        
        d_row.extend(right_cells)
        rows.append(d_row)
        rows.append([GridCell("", cell_type='separator')])

    # ---------------------------------------------------------
    # 1. Right Dependents
    
    # ---------------------------------------------------------
    # 1. Right Dependents
    # ---------------------------------------------------------
    for tot in [4, 3, 2, 1]:
        row_cells = []
        row_cells.append(GridCell(f"R tot={tot}", cell_type='label'))
        row_cells.append(GridCell("")) # Mirror Extra Empty
        
        # Padding
        for _ in range(7): row_cells.append(GridCell(""))
        
        # Center "V"
        row_cells.append(GridCell("V", cell_type='label'))
        
        prev_val = None
        
        for pos in range(1, tot + 1):
             key = f'right_{pos}_totright_{tot}'
             val = position_averages.get(key, float('nan'))
             
             # Horizontal Factor logic
             if pos > 1 and show_horizontal_factors:
                 fac_cell = GridCell("", cell_type='factor')
                 rich_segments = []
                 text_parts = []
                 
                 # 1. Factor
                 factor_val = None
                 key_b = f'right_{pos}_totright_{tot}'
                 key_a = f'right_{pos-1}_totright_{tot}'
                 fac_key = f'factor_{key_b}_vs_{key_a}'
                 geo_factor = position_averages.get(fac_key)

                 if geo_factor is not None:
                     factor_val = geo_factor
                 elif prev_val is not None and val is not None and val != 0 and val == val:
                     factor_val = val / prev_val
                 if factor_val is not None:
                     is_outward = get_growth_direction('right', arrow_direction) == 'outward'
                     val_factor_display = factor_val if is_outward else (1.0 / factor_val if factor_val != 0 else 1.0)
                     sym = get_arrow('right', 'horizontal', arrow_direction)
                     f_str = f"×{val_factor_display:.2f}{sym}"
                     
                     # Color Rule: Red if factor < 1
                     f_color = COLOR_RED if val_factor_display < 1.0 else COLOR_GREY
                     rich_segments.append((f_str, f_color, True))
                     text_parts.append(f_str)

                 # 2. Triple
                 if show_ordering_triples and ordering_stats:
                     pair_idx = pos - 2
                     o_key = ('right', tot, pair_idx)
                     o_data = ordering_stats.get(o_key)
                     if o_data:
                        tot_count = o_data['lt'] + o_data['eq'] + o_data['gt']
                        if tot_count > 0:
                            lt = o_data['lt'] / tot_count * 100
                            eq = o_data['eq'] / tot_count * 100
                            gt = o_data['gt'] / tot_count * 100
                            
                            # Build text for TSV
                            t_str_full = f"(<{lt:.0f}={eq:.0f}>{gt:.0f})"
                            text_parts.append(t_str_full)
                            
                            # Build Rich Text Segments
                            # Prefix: " (<" + lt + "=" + eq + ">"
                            # User wants: (<31=35>34) where 34 is Red if gt > lt
                            
                            # Add space separator if factor exists
                            if rich_segments:
                                rich_segments.append((" ", COLOR_GREY, False))
                                
                            prefix = f"(<{lt:.0f}={eq:.0f}>"
                            rich_segments.append((prefix, COLOR_GREY, False))
                            
                            # GT part
                            gt_str = f"{gt:.0f}"
                            gt_color = COLOR_RED if gt > lt else COLOR_GREY
                            gt_bold = True if gt > lt else False
                            rich_segments.append((gt_str, gt_color, gt_bold))
                            
                            # Suffix
                            rich_segments.append((")", COLOR_GREY, False))

                 if text_parts:
                     fac_cell.text = " ".join(text_parts)
                     fac_cell.rich_text = rich_segments
                     
                 row_cells.append(fac_cell)
             
             # Value Cell (Numeric)
             val_cell = GridCell("", value=val, cell_type='value')
             if isinstance(val, float) and val == val: # not nan
                 val_cell.text = f"{val:.3f}"
                 # Ensure value is stored for Excel
                 val_cell.value = float(val) 
             else:
                 val_cell.text = "N/A"
                 val_cell.value = None
             row_cells.append(val_cell)
             
             prev_val = val
             
        # Add Extra Empty Column at end of Grid (before comment)
        row_cells.append(GridCell(""))
                     
        # Add Gap for Extra Col in Text Output
        # parts.append(E_FAC) # Assuming FAC_WIDTH matches
        
        if show_row_averages:
             avg_key = f'average_totright_{tot}'
             row_avg = position_averages.get(avg_key)
             comment_str = ""
             if row_avg is not None:
                 comment_str = f"[GM: {row_avg:.3f}]"
             
             # Add N if available
             if ordering_stats:
                 # Try new explicit total
                 t_key = ('right', tot, 'total')
                 n_count = ordering_stats.get(t_key)
                 
                 # Fallback to pair sum if total not found (legacy data)
                 if n_count is None and tot >= 2:
                     o_key = ('right', tot, 0)
                     o_data = ordering_stats.get(o_key)
                     if o_data:
                         n_count = o_data['lt'] + o_data['eq'] + o_data['gt']
                 
                 if n_count is not None:
                     comment_str += f" [N={n_count}]"

             # Add Slope (Tot >= 2)
             if tot >= 2:
                 y_vals = []
                 valid_slope = True
                 for pos in range(1, tot + 1):
                     k = f'right_{pos}_totright_{tot}'
                     v = position_averages.get(k)
                     if v is None: 
                         valid_slope = False
                         break
                     y_vals.append(v)
                 
                 if valid_slope and len(y_vals) == tot:
                     try:
                         slope, _ = np.polyfit(np.arange(tot), y_vals, 1)
                         comment_str += f" [Slope: {slope:+.2f}]"
                     except:
                         pass
             if comment_str:
                 # Pad the row to ensure comment is in the last column
                 # Target Len: Label + Mirror + Left(7) + V + Right(7) + Extra(1) = 1 + 1 + 7 + 1 + 7 + 1 = 18.
                 current_len = len(row_cells)
                 needed = 18 - current_len
                 if needed > 0:
                     for _ in range(needed):
                         row_cells.append(GridCell(""))
                         
                 row_cells.append(GridCell(comment_str, cell_type='comment'))
                 
                 # Ensure we didn't overshoot (unlikely with loop logic but safe to check)
                 # Actually if show_horizontal_factors=False, width is less.
                 # But Header assumes max width?
                 # Header generation uses check.
                 # If show_horizontal_factors=False, Max R cols is 4.
                 # My Header Generation makes 4 cols + 3 empty factors if False? No.
                 # Header generation: `if pos < 4 and show_horizontal_factors: append factor`.
                 # So if False, Header R-block is 4 cells.
                 # Target length: 1 + 7 + 1 + 4 = 13.
                 # If True, Target: 1 + 7 + 1 + 7 = 16.
                 
        rows.append(row_cells)
        
        # Diagonals (Right)
        if tot > 1 and show_diagonal_factors:
            diag_row = [GridCell(f"Diag R{tot}-{tot-1}", cell_type='label')]
            diag_row.append(GridCell("")) # Mirror Extra
            for _ in range(7): diag_row.append(GridCell("")) 
            diag_row.append(GridCell("")) 
            
            for pos in range(1, tot):
                diag_row.append(GridCell("")) 
                if show_horizontal_factors:
                    src_key = f'right_{pos}_totright_{tot-1}'
                    tgt_key = f'right_{pos+1}_totright_{tot}'
                    src_val = position_averages.get(src_key)
                    tgt_val = position_averages.get(tgt_key)
                    
                    diag_cell = GridCell("", cell_type='factor')
                    
                    factor = None
                    is_outward = get_growth_direction('right', arrow_direction) == 'outward'
                    if geo_factor is not None:
                        factor_val = geo_factor if is_outward else (1.0/geo_factor if geo_factor != 0 else 1.0)
                    elif src_val and tgt_val:
                        factor_val = tgt_val / src_val if is_outward else src_val / tgt_val
                        
                    if factor_val is not None:
                        sym = get_arrow('right', 'diagonal', arrow_direction)
                        diag_str = f"×{factor_val:.2f} {sym}"
                        diag_cell.text = diag_str
                        # Rich text for diagonal
                        d_color = COLOR_RED if factor_val < 1.0 else COLOR_GREY
                        diag_cell.rich_text = [(diag_str, d_color, True)]
                    diag_row.append(diag_cell)
            rows.append(diag_row)

    rows.append([GridCell("", cell_type='separator')])
    
    # ---------------------------------------------------------
    # 3. Left Marginal Means (Bottom)
    # ---------------------------------------------------------
    if show_marginal_means:
        calc_gm = calc_marginal_gm # alias
        rows.append([GridCell("", cell_type='separator')])
        
        # -----------------------------------------------
        # 3A. Vertical Means (Left Row 1)
        # -----------------------------------------------
        # Col L_P: GM(Sizes in Column L_P).
        # Col F_P (ValIdx + 1): GM(H-Factor) and GM(D-Factor).
        # H-Factor L_P -> L_{P-1} (Inner).
        # D-Factor L_P(T) -> L_{P-1}(T)? 
        # Left D-Factors connect Tgt/Src.
        # My Left Loop Logic: Src (P+1) -> Tgt (P)? No.
        # logic: key_b (L_P, T) / key_a (L_{P-1}, T) for H-Factor (Inner/Outer).
        # logic: factor_ L_P_T_vs_L_{P+1}_{T+1} for D-Factor.
        # So D-Factor connects L_{P+1}(T+1) -> L_P(T).
        # i.e. Outer/HighTot -> Inner/LowTot.
        # Wait, direction is usually Outer -> Inner?
        # L4 -> L3?
        # The grid display puts Factor between L_{P+1} and L_P?
        # My grid loop iterates P=1..Tot.
        # Displays `Val(L_P)` then `Factor(L_P / L_{P-1})`.
        # So Factor is towards V.
        # Diagonal Factor `diag_cells[fac_idx]` is placed...
        # Loop diagonal: `tgt=L_P(T)`, `src=L_{P+1}(T+1)`. (Tgt is inner).
        # `fac_idx` is left of `pos`.
        # So Factor connects `Col L_{P+1}` and `Col L_P`.
        # It sits between them.
        
        # Let's align with columns:
        # Col L4. Col F43. Col L3.
        # F43 connects L4 -> L3.
        # So we want means of factors L4->L3.
        # H-Factor: L4(T) -> L3(T).
        # D-Factor: ??
        # In Grid Diag Row: `src=L_{pos+1}, tgt=L_{pos}`.
        # If `pos=3` (Target L3). Src L4.
        # Factor connects L4(T+1) -> L3(T).
        # This sits at `fac_idx` for `pos=3`. i.e. between L3 and L4.
        # So yes.
        # At transition P (Outer) -> P-1 (Inner):
        #   H-Factor: L_P(T) -> L_{P-1}(T).
        #   D-Factor: L_P(T+1) -> L_{P-1}(T).
        
        # So for `pos` in 4 down to 2?
        # Loop `p` from 2 to 4. Transition `p -> p-1`.
        
        # -----------------------------------------------
        # 3A. Vertical Means (Left Row 1)
        # -----------------------------------------------
        
        # Initialize Left Variables
        l_vert_size = {}
        l_vert_h = {}
        l_vert_d = {}
        
        for pos in range(1, 5):
            # Size
            keys = [f'left_{pos}_totleft_{tot}' for tot in range(pos, 5)]
            l_vert_size[pos] = calc_gm(keys)
            
            if pos > 1:
                # H-Factor: L_P(T) -> L_{P-1}(T).
                h_keys = []
                for tot in range(pos, 5):
                    h_keys.append(f'factor_left_{pos}_totleft_{tot}_vs_left_{pos-1}_totleft_{tot}')
                l_vert_h[pos] = calc_gm(h_keys)
                
                # D-Factor: L_{P-1}(3) vs L_{P}(4) (e.g. Diag 0)
                # Key: factor_left_{pos-1}_totleft_3_vs_left_{pos}_totleft_4
                d_key = f'factor_left_{pos-1}_totleft_3_vs_left_{pos}_totleft_4'
                val = position_averages.get(d_key)
                if val:
                    l_vert_d[pos] = 1.0 / val # Invert for Growth Outward

        # Label

        
        # 3B. Diagonal Means (Left Row 2)
        # -----------------------------------------------
        # Map Diagonals.
        # L4 (Outer) -> Diag 0 (Last X).
        # L1 (Inner) -> Diag 3 (First X).
        l_diag_size = {}
        l_diag_h = {}
        l_diag_d = {} # Diagonal factors for the "Diagonal Rows"
        
        # Size Means (per Diag K)
        # K=0 (Last): L1T1..L4T4.
        # K=3 (First): L1T4.
        # Matches Right side structure.
        raw_d_size = {}
        for k in range(4):
            keys = []
            for pos in range(1, 5):
                tot = pos + k
                if tot <= 4:
                    keys.append(f'left_{pos}_totleft_{tot}')
            raw_d_size[k] = calc_gm(keys)

        for pos in range(1, 5):
            k = 4 - pos
            l_diag_size[pos] = raw_d_size.get(k)
            
            # Factors between `L_P` (Diag K) and `L_{P-1}` (Diag K+1/Inner).
            # We want factors connecting them for the "Diagonal Row".
            # H-Factor: Connects Diag K -> Diag K+1?
            #   Transition P -> P-1.
            #   P corresponds to Diag K=4-P.
            #   P-1 corresponds to Diag J=4-(P-1) = 5-P = K+1.
            #   So transition Outer(P) -> Inner(P-1).
            #   H-Factor L_P(T) -> L_{P-1}(T).
            #   T-P = K => T = P+K = P + 4-P = 4.
            #   So T=4.
            #   Mean of H-Factors at T=4.
            if pos > 1:
                # H-Factor: L_P(4) -> L_{P-1}(4).
                h_key = f'factor_left_{pos}_totleft_4_vs_left_{pos-1}_totleft_4'
                l_diag_h[pos] = position_averages.get(h_key)
                
                # D-Factor key: L_{P-1}(T=3) vs L_{P}(T=4).
                d_key = f'factor_left_{pos-1}_totleft_3_vs_left_{pos}_totleft_4'
                val = position_averages.get(d_key)
                if val:
                    l_diag_d[pos] = 1.0 / val # Invert for Growth Outward
                
        # Construct Rows
        
        # Row 1: [M Diag Left]
        ld_row = [GridCell("M Diag Left", cell_type='label', rich_text=[("M Diag Left", COLOR_ORANGE, True)])]
        
        # mirror_extra_cell (Longest Diag Len 3)
        gm_d3 = raw_d_size.get(0)
        
        keys_fac_d3 = []
        for pos in range(2, 5):
            keys_fac_d3.append(f'factor_left_{pos}_totleft_{pos}_vs_left_{pos-1}_totleft_{pos-1}')
        gm_fac_d3 = calc_gm(keys_fac_d3)
        
        def make_fac_left(fac, sz):
            if fac is None: return GridCell("")
            val = fac if is_outward else (1.0 / fac if fac != 0 else 1.0)
            sym = get_arrow('left', 'diagonal', arrow_direction)
            s = f"{sz:.2f} ×{val:.2f}{sym}" if sz is not None else f"×{val:.2f}{sym}"
            return GridCell(s, cell_type='factor', rich_text=[(s, COLOR_ORANGE, False)])
            
        mirror_extra_cell = make_fac_left(gm_fac_d3, gm_d3)
        ld_row.append(mirror_extra_cell)
        
        left_grid_d = [GridCell("") for _ in range(7)]
        for pos in range(1, 5):
            gm = l_diag_size.get(pos)
            if gm is not None:
                if show_horizontal_factors:
                    idx = 6 - (pos - 1) * 2
                else:
                    idx = 3 - (pos - 1)
                if 0 <= idx < 7:
                     txt = f"{gm:.3f}"
                     left_grid_d[idx] = GridCell(txt, value=float(gm), cell_type='value', rich_text=[(txt, COLOR_ORANGE, True)])
                     
            if pos > 1:
                 h_gm = l_diag_h.get(pos)
                 d_gm = l_diag_d.get(pos) # Already inverted in calculation loop
                 
                 parts = []
                 rich = []
                 if h_gm is not None:
                    val = h_gm if is_outward else (1.0 / h_gm if h_gm != 0 else 1.0)
                    sym = get_arrow('left', 'horizontal', arrow_direction)
                    s = f"×{val:.2f}{sym}"
                    parts.append(s)
                    rich.append((s, COLOR_ORANGE, False))
                 if d_gm is not None:
                    val = d_gm if is_outward else (1.0 / d_gm if d_gm != 0 else 1.0)
                    sym = get_arrow('left', 'diagonal', arrow_direction)
                    s = f"×{val:.2f}{sym}" # Arrow bottom-left
                    parts.append(s)
                    if h_gm is not None: 
                        parts[-2] += " "
                        rich.insert(-1, (" ", COLOR_ORANGE, False))
                    rich.append((s, COLOR_ORANGE, False))
                 
                 if parts and show_horizontal_factors:
                     idx = 6 - (pos - 1) * 2
                     fac_idx = idx + 1
                     if fac_idx < 7:
                         left_grid_d[fac_idx] = GridCell(" ".join(parts), cell_type='factor', rich_text=rich)

        ld_row.extend(left_grid_d)
        # Removed V column
        ld_row.append(GridCell("")) # Extra column for alignment
        for _ in range(8): ld_row.append(GridCell("")) # Right Side Empty
        
        rows.append(ld_row)

        
        # Row 2: [M Vert Left]
        lv_row = [GridCell("M Vert Left", cell_type='label', rich_text=[("M Vert Left", COLOR_BLUE, True)])]
        lv_row.append(GridCell("")) # Mirror Extra
        
        left_grid_v = [GridCell("") for _ in range(7)]
        for pos in range(1, 5):
            gm = l_vert_size.get(pos)
            if gm is not None:
                # Index logic
                if show_horizontal_factors:
                    idx = 6 - (pos - 1) * 2
                else:
                    idx = 3 - (pos - 1)
                
                if 0 <= idx < 7:
                    txt = f"{gm:.3f}"
                    left_grid_v[idx] = GridCell(txt, value=float(gm), cell_type='value', rich_text=[(txt, COLOR_BLUE, True)])
                    
            # Factors (between Pos and Pos-1)
            # Display at `fac_idx` corresponding to Pos-1? No.
            # Grid: L4 F43 L3 ...
            # F43 is after L4.
            # If `idx` is L4 (0). F43 is `idx+1` (1).
            # This corresponds to factor L4->L3. (Pos=4).
            if pos > 1:
                 h_gm = l_vert_h.get(pos)
                 d_gm = l_vert_d.get(pos)
                 
                 parts = []
                 rich = []
                 # Arrow pointing Left (Outward)
                 if h_gm is not None:
                    # Invert if stored as Inner/Outer?
                    # Stored `L_P / L_{P-1}` (Outer/Inner). Growth is > 1. Correct.
                    is_outward = get_growth_direction('left', arrow_direction) == 'outward'
                    val = h_gm if is_outward else (1.0 / h_gm if h_gm != 0 else 1.0)
                    sym = get_arrow('left', 'horizontal', arrow_direction)
                    s = f"×{val:.2f}{sym}"
                    parts.append(s)
                    rich.append((s, COLOR_BLUE, False))
                 
                 if d_gm is not None:
                    # Stored `L_{P-1} / L_P` (Inner/Outer) for Diagonals??
                    # My logic above: D-Factor is `Tgt(Inner)/Src(Outer)`.
                    # So need invert.
                    # d_val = 1.0 / d_gm # d_gm is already inverted
                    is_outward = get_growth_direction('left', arrow_direction) == 'outward'
                    val = d_gm if is_outward else (1.0 / d_gm if d_gm != 0 else 1.0)
                    sym = get_arrow('left', 'diagonal', arrow_direction)
                    s = f"×{val:.2f}{sym}"
                    parts.append(s)
                    if h_gm is not None: 
                        parts[-2] += " "
                        rich.insert(-1, (" ", COLOR_BLUE, False))
                    rich.append((s, COLOR_BLUE, False))
                    
                 if parts and show_horizontal_factors:
                     idx = 6 - (pos - 1) * 2 # Index of L_Pos
                     fac_idx = idx + 1      # Index of F_{P, P-1}
                     if fac_idx < 7:
                         left_grid_v[fac_idx] = GridCell(" ".join(parts), cell_type='factor', rich_text=rich)

        lv_row.extend(left_grid_v)
        # Removed V column
        lv_row.append(GridCell("")) # Extra column for alignment
        for _ in range(8): lv_row.append(GridCell("")) # Right Side Empty
        rows.append(lv_row)
        
    # ---------------------------------------------------------
    # 1.5. XVX Configuration (L1 V R1)
    # ---------------------------------------------------------
    l_xvx = position_averages.get('xvx_left_1')
    r_xvx = position_averages.get('xvx_right_1')
    
    if l_xvx is not None or r_xvx is not None:
        row_cells = []
        row_cells.append(GridCell("X V X", cell_type='label'))
        row_cells.append(GridCell("")) # Mirror Extra
        
        # Left Grid: L1 is at index 6 (last slot)
        left_grid = [GridCell("") for _ in range(7)]
        if l_xvx is not None:
            left_grid[6] = GridCell(f"{l_xvx:.3f}", value=float(l_xvx), cell_type='value')
        else:
            left_grid[6] = GridCell("N/A", value=None)
            
        row_cells.extend(left_grid)
        row_cells.append(GridCell("V", cell_type='label'))
        
        # Right Grid: R1 is at first slot?
        # My header for Right is R1, F, R2...
        # So R1 is appended immediately.
        if r_xvx is not None:
             row_cells.append(GridCell(f"{r_xvx:.3f}", value=float(r_xvx), cell_type='value'))
        else:
             row_cells.append(GridCell("N/A", value=None))
        
        # Factor L1 -> R1 (Cross-verb)
        if show_horizontal_factors and l_xvx and r_xvx and l_xvx != 0:
            fac_cell = GridCell("", cell_type='factor')
            rich_segments = []
            text_parts = []
            
            # Factor
            factor_val = None
            fac_key = 'factor_xvx_right_1_vs_xvx_left_1'
            geo_factor = position_averages.get(fac_key)
            
            if geo_factor is not None:
                factor_val = geo_factor
            elif l_xvx is not None and l_xvx != 0:
                 factor_val = r_xvx / l_xvx
                 
            if factor_val is not None:
                # Apply inward logic
                val = factor_val if is_outward else (1.0 / factor_val if factor_val != 0 else 1.0)
                # XVX is L -> R. Outward is →. Inward is ←.
                sym = get_arrow('right', 'horizontal', arrow_direction) # Assuming XVX is like a rightward horizontal factor
                f_str = f"×{val:.2f}{sym}"
                f_color = COLOR_RED if val < 1.0 else COLOR_GREY
                rich_segments.append((f_str, f_color, True))
                text_parts.append(f_str)
            
            # Ordering Stats
            if show_ordering_triples and ordering_stats:
                o_key = ('xvx', 2, 0) # L1 vs R1 pair
                o_data = ordering_stats.get(o_key)
                if o_data:
                    tot_count = o_data['lt'] + o_data['eq'] + o_data['gt']
                    if tot_count > 0:
                        lt = o_data['lt'] / tot_count * 100
                        eq = o_data['eq'] / tot_count * 100
                        gt = o_data['gt'] / tot_count * 100
                        trip_str = f"(<{lt:.0f}={eq:.0f}>{gt:.0f})"
                        text_parts.append(trip_str)
                        
                        if rich_segments: rich_segments.append((" ", COLOR_GREY, False))
                        rich_segments.append((f"(<{lt:.0f}={eq:.0f}>", COLOR_GREY, False))
                        gt_color = COLOR_RED if gt > lt else COLOR_GREY
                        gt_bold = True if gt > lt else False
                        rich_segments.append((f"{gt:.0f}", gt_color, gt_bold))
                        rich_segments.append((")", COLOR_GREY, False))
                        
            if text_parts:
                fac_cell.text = " ".join(text_parts)
                fac_cell.rich_text = rich_segments
                row_cells.append(fac_cell)
            else:
                row_cells.append(GridCell(""))
        else:
            if show_horizontal_factors:
                row_cells.append(GridCell(""))
        
        # Padding for rest of Right Side
        # Total Right side cells target: 7.
        # We added R1 (1) + Fac (1). Total 2.
        # Need 5 more.
        # If fac didn't add (False), we added 1. Need 3? (4 total if no fac).
        current_data_len = len(row_cells) - (1 + 7 + 1) # Label, LeftGrid, V
        target = 7 if show_horizontal_factors else 4
        while current_data_len < target:
            row_cells.append(GridCell(""))
            current_data_len += 1
            
        # Comment: N / Slope
        comment_str = ""
        # N
        if ordering_stats:
             n_key = ('xvx', 2, 'total')
             n_val = ordering_stats.get(n_key)
             # Fallback sum
             if n_val is None:
                  o_key = ('xvx', 2, 0)
                  d = ordering_stats.get(o_key)
                  if d: n_val = d['lt'] + d['eq'] + d['gt']
             
             if n_val is not None:
                  comment_str += f" [N={n_val}]"
        
        # Slope? L1 and R1.
        # x coordinates: L1 at -1? R1 at +1?
        # Or visual 1, 2? L1->1, R1->2?
        # "Slope" usually implies dependent vs position.
        # Here we have L1 and R1. 
        # Maybe slope L1->R1? (y2-y1)/(x2-x1).
        # Diff = R1 - L1.
        # If displayed as slope, sure.
        if l_xvx is not None and r_xvx is not None:
             slope = r_xvx - l_xvx # Change over 1 unit? Or 2? 
             # L1 is pos 1 Left. R1 is pos 1 Right.
             # Distance? 
             # Let's maybe skip Slope for XVX unless requested.
             pass

        if comment_str:
             row_cells.append(GridCell(comment_str, cell_type='comment'))
        
        rows.append(row_cells)
        rows.append([GridCell("", cell_type='separator')])
    
    # ---------------------------------------------------------
    # 2. Left Dependents
    # ---------------------------------------------------------
    for tot in [1, 2, 3, 4]:
        row_cells = []
        row_cells.append(GridCell(f"L tot={tot}", cell_type='label'))
        row_cells.append(GridCell("")) # Mirror Extra
        
        left_grid = [GridCell("") for _ in range(7)]
        prev_val_inner = None
        
        for pos in range(1, tot + 1):
             key = f'left_{pos}_totleft_{tot}'
             val = position_averages.get(key, None)
             
             if show_horizontal_factors:
                 val_idx = 6 - (pos - 1) * 2
             else:
                  val_idx = 3 - (pos - 1)
              
              # If we added marginal means on Left, indices must align.
              # Left headers: L4 down to L1. 
              # L1 is at index 6 (if factors) or 3 (no factors) in the Left Grid List.
              
              # Value
             val_cell = GridCell("", value=val, cell_type='value')
             if val is not None: 
                 val_cell.text = f"{val:.3f}"
                 val_cell.value = float(val)
             else:
                 val_cell.text = "N/A"
             if val_idx >= 0 and val_idx < 7:
                 left_grid[val_idx] = val_cell
             
             # Factor (to the right of value)
             fac_idx = val_idx + 1
             if pos > 1 and show_horizontal_factors and fac_idx < 7:
                 fac_cell = GridCell("", cell_type='factor')
                 rich_segments = []
                 text_parts = []
                 
                 # 1. Factor
                 factor_val = None
                 key_b = f'left_{pos}_totleft_{tot}'
                 key_a = f'left_{pos-1}_totleft_{tot}'
                 fac_key = f'factor_{key_b}_vs_{key_a}'
                 geo_factor = position_averages.get(fac_key)
                 
                 if geo_factor is not None:
                     factor_val = geo_factor
                 elif prev_val_inner is not None and val is not None and prev_val_inner != 0:
                     factor_val = val / prev_val_inner
                     
                 if factor_val is not None:
                     # Apply inward logic
                     val_factor_display = factor_val if is_outward else (1.0 / factor_val if factor_val != 0 else 1.0)
                     sym = get_arrow('left', 'horizontal', arrow_direction)
                     f_str = f"×{val_factor_display:.2f}{sym}"
                     f_color = COLOR_RED if val_factor_display < 1.0 else COLOR_GREY
                     rich_segments.append((f_str, f_color, True))
                     text_parts.append(f_str)

                 # 2. Triple
                 if show_ordering_triples and ordering_stats:
                     pair_idx = tot - pos
                     o_key = ('left', tot, pair_idx)
                     o_data = ordering_stats.get(o_key)
                     if o_data:
                        tot_count = o_data['lt'] + o_data['eq'] + o_data['gt']
                        if tot_count > 0:
                            lt = o_data['lt'] / tot_count * 100
                            eq = o_data['eq'] / tot_count * 100
                            gt = o_data['gt'] / tot_count * 100
                            
                            t_str_full = f"(<{lt:.0f}={eq:.0f}>{gt:.0f})"
                            text_parts.append(t_str_full)

                            if rich_segments:
                                rich_segments.append((" ", COLOR_GREY, False))
                                
                            prefix = f"(<{lt:.0f}={eq:.0f}>"
                            rich_segments.append((prefix, COLOR_GREY, False))
                            
                            gt_str = f"{gt:.0f}"
                            gt_color = COLOR_RED if gt > lt else COLOR_GREY
                            gt_bold = True if gt > lt else False
                            rich_segments.append((gt_str, gt_color, gt_bold))
                            
                            rich_segments.append((")", COLOR_GREY, False))

                 if text_parts:
                     fac_cell.text = " ".join(text_parts)
                     fac_cell.rich_text = rich_segments
                     
                 left_grid[fac_idx] = fac_cell
             
             prev_val_inner = val
             
        row_cells.extend(left_grid)
        row_cells.append(GridCell("V", cell_type='label'))
        
        # Row Average
        if show_row_averages:
             avg_key = f'average_totleft_{tot}'
             row_avg = position_averages.get(avg_key)
             comment_str = ""
             if row_avg is not None:
                 comment_str = f"[GM: {row_avg:.3f}]"
             
             # Add N if available
             if ordering_stats:
                 t_key = ('left', tot, 'total')
                 n_count = ordering_stats.get(t_key)
                 
                 if n_count is None and tot >= 2:
                     o_key = ('left', tot, 0)
                     o_data = ordering_stats.get(o_key)
                     if o_data:
                         n_count = o_data['lt'] + o_data['eq'] + o_data['gt']
                 
                 if n_count is not None:
                     comment_str += f" [N={n_count}]"
                     
             # Add Slope (Tot >= 2)
             if tot >= 2:
                 y_vals = []
                 valid_slope = True
                 # Visual Order: Outer (L_tot) -> Inner (L1)
                 # Keys are left_{pos}_... where pos 1 is L1.
                 # So iterate pos from tot down to 1.
                 for pos in range(tot, 0, -1):
                     k = f'left_{pos}_totleft_{tot}'
                     v = position_averages.get(k)
                     if v is None: 
                         valid_slope = False
                         break
                     y_vals.append(v)
                     
                 if valid_slope and len(y_vals) == tot:
                     try:
                         slope, _ = np.polyfit(np.arange(tot), y_vals, 1)
                         comment_str += f" [Slope: {slope:+.2f}]"
                     except:
                         pass
                     
        # Left Rows need padding for the Right side gap
        # 1. Label (1) - already added
        # 2. LeftGrid (7) - already added/extended
        # 3. V (1) - already added
        
        # 4. Right Side Padding. 
        #    If factors: 7 cells. If no factors: 4 cells.
        r_pad_len = 7 if show_horizontal_factors else 4
        for _ in range(r_pad_len):
            row_cells.append(GridCell(""))
        
        # Add Extra Column Gap
        row_cells.append(GridCell(""))
             
        # 5. Comment (Avg | N | Slope)
        if comment_str:
             row_cells.append(GridCell(comment_str, cell_type='comment'))
        
        rows.append(row_cells)
        
        # Diagonals (Left)
        # Logic matches print function: diagonals between Tot and Tot+1
        if tot < 4 and show_diagonal_factors and show_horizontal_factors:
            diag_row = [GridCell(f"Diag L{tot}-{tot+1}", cell_type='label')]
            diag_row.append(GridCell("")) # Mirror Extra
            diag_grid = [GridCell("") for _ in range(7)]
            
            for pos in range(1, tot + 1):
                tgt_key = f'left_{pos}_totleft_{tot}'
                src_key = f'left_{pos+1}_totleft_{tot+1}'
                tgt_val = position_averages.get(tgt_key)
                src_val = position_averages.get(src_key)
                
                if pos + 1 > 4: continue
                
                # Slot left of pos. 
                # pos 1 (L1) is at index 6. Slot left is 5.
                # pos 2 (L2) is at index 4. Slot left is 3.
                # Formula: (6 - (pos - 1) * 2) - 1
                fac_idx = (6 - (pos - 1) * 2) - 1
                
                factor_val = None
                fac_key = f'factor_{tgt_key}_vs_{src_key}'
                geo_factor = position_averages.get(fac_key)

                if geo_factor is not None:
                     is_outward = get_growth_direction('left', arrow_direction) == 'outward'
                     factor_val = geo_factor if is_outward else (1.0 / geo_factor if geo_factor != 0 else 1.0)
                elif src_val and tgt_val:
                     is_outward = get_growth_direction('left', arrow_direction) == 'outward'
                     factor_val = tgt_val / src_val if is_outward else src_val / tgt_val
                         
                if factor_val is not None:
                     sym = get_arrow('left', 'diagonal', arrow_direction)
                     diag_str = f"×{factor_val:.2f} {sym}"
                     
                     diag_cell = GridCell(diag_str, cell_type='factor')
                     d_color = COLOR_RED if factor_val < 1.0 else COLOR_GREY
                     diag_cell.rich_text = [(diag_str, d_color, True)]
                     
                     if 0 <= fac_idx < 7:
                         diag_grid[fac_idx] = diag_cell
                        
            diag_row.extend(diag_grid)
            diag_row.append(GridCell("")) # V column
            rows.append(diag_row)

    return rows

def _format_verb_centered_table_legacy(position_averages, 
                               show_horizontal_factors=False, 
                               show_diagonal_factors=False,
                               arrow_direction='diverging',
                               disorder_percentages=None,
                               ordering_stats=None,
                               show_ordering_triples=False,
                               show_row_averages=False,
                               show_marginal_means=True,
                               save_tsv=True,
                               output_dir='data',
                               filename=None):
    """
    LEGACY IMPLEMENTATION - Format the computed averages into a readable string table.
    Optionally saves the table to a TSV file.
    
    Parameters
    ----------
    position_averages : dict
        Dictionary of values.
    show_horizontal_factors : bool
        If True, adds columns with growth factors between positions.
    show_diagonal_factors : bool
        If True, adds rows with diagonal growth factors.
    arrow_direction : str
        'diverging': Arrows point outward from verb (Right->R, Left->L).
        'rightwards': Arrows point generally Top-Right (Right->R, Left->R).
    disorder_percentages : dict, optional
        Dictionary: (side, tot) -> percentage of disordered languages.
        If provided, adds disorder percentage to each row.
    ordering_stats : dict, optional
        Dictionary from get_ordering_stats with granular counts.
    show_ordering_triples : bool
        If True, replace horizontal factors with (lt, eq, gt) percentages.
        Requires ordering_stats.
    show_row_averages : bool
        If True, append average constituent size for the row.
    save_tsv : bool
        If True, save the table to a file.
    output_dir : str
        Directory to save output.
    filename : str
        Filename for TSV output. If None, auto-generated from options.
    """
    lines = []
    is_outward = arrow_direction in ['diverging', 'rightwards', 'outward']
    lines.append("=" * 120)
    lines.append("VERB-CENTERED CONSTITUENT SIZE TABLE")
    if show_horizontal_factors or show_diagonal_factors:
        lines.append("WITH GROWTH FACTORS")
    if show_ordering_triples:
        lines.append("WITH ORDERING TRIPLES ( < % | = % | > % )")
    lines.append("=" * 120)
    lines.append("")
    
    # Layout constants
    VAL_WIDTH = 8   
    FAC_WIDTH = 22 if show_ordering_triples else (12 if show_horizontal_factors else 0)
    E_VAL = " " * VAL_WIDTH
    E_FAC = " " * FAC_WIDTH
    HALF_WIDTH = 4 * VAL_WIDTH + 3 * FAC_WIDTH
    LEFT_PAD = " " * HALF_WIDTH
    
    # Header V
    HALF_WIDTH = 4 * VAL_WIDTH + 3 * FAC_WIDTH
    
    # Text Header Row
    # Ident matching "R tot=X: " (9 chars)
    header_parts = [" " * 9]
    header_parts.append(" " * FAC_WIDTH) # Mirrored Extra Column
    
    # Left Headers (L4 .. L1)
    l_headers = []
    for i in range(4, 0, -1):
        l_headers.append(f"L{i}".center(VAL_WIDTH))
        if i > 1 and (show_horizontal_factors or show_diagonal_factors):
            # Use FAC_WIDTH spaces or a label if desired
            l_headers.append(E_FAC)
    header_parts.append("".join(l_headers))
    
    # Center V
    header_parts.append(" V ") # Matches ' V ' (3 chars) approx logic
     # Actually data rows use ' V' (2 chars) near R side or ' V' near L side.
     # L-rows: left_str + ' V'.
     # R-rows: LEFT_PAD + ' V' + right_str_content.
    
    # Right Headers (R1 .. R4)
    r_headers = []
    for i in range(1, 5):
        r_headers.append(f"R{i}".center(VAL_WIDTH))
        if i < 4 and (show_horizontal_factors or show_diagonal_factors):
            r_headers.append(E_FAC)
    
    # Extra Empty Column before Stats
    r_headers.append(" " * FAC_WIDTH) # Col 8 (Index 7)
                                      # Wait. Indices: 0(R1), 1(F12), 2(R2), 3(F23), 4(R3), 5(F34), 6(R4).
                                      # New Col is Index 7.
                                      # Use FAC_WIDTH for the empty column? Or VAL_WIDTH?
                                      # User: "add an extra empty column". Factor column size is appropriate for factors.
                                      # M Diag Right will place factor there.
    
    header_parts.append("".join(r_headers))
    
    if show_row_averages:
        header_parts.append("  [GM | N | Slope]")
        
    lines.append("".join(header_parts)) 



    tsv_rows = []
    header_cols = ["Row", "", "L4", "F43", "L3", "F32", "L2", "F21", "L1", "V", "R1", "F12", "R2", "F23", "R3", "F34", "R4", ""]
    if show_row_averages:
        header_cols.append("Stats") # GM | N | Slope
    tsv_rows.append(header_cols)

    # Helper
    def calc_marginal_gm_txt(keys):
        vals = []
        for k in keys:
            v = position_averages.get(k)
            if v is not None and v > 0:
                vals.append(v)
        if not vals: return None
        return np.exp(np.mean(np.log(vals)))

    # ---------------------------------------------------------
    # 0. Right Marginal Means (Top)
    # ---------------------------------------------------------
    if show_marginal_means:
        calc_gm = calc_marginal_gm_txt # alias
        
        # Padding calculation (Empty Left Grid)
        pad_cells = []
        pad_cells.append(" " * VAL_WIDTH) # L4
        if show_horizontal_factors: pad_cells.append(" " * FAC_WIDTH) # F43
        pad_cells.append(" " * VAL_WIDTH) # L3
        if show_horizontal_factors: pad_cells.append(" " * FAC_WIDTH) # F32
        pad_cells.append(" " * VAL_WIDTH) # L2
        if show_horizontal_factors: pad_cells.append(" " * FAC_WIDTH) # F21
        pad_cells.append(" " * VAL_WIDTH) # L1
        
        left_padding_str = "".join(pad_cells)
        # Header indent is 9 spaces. Top Rows must match.
        indent_str = " " * 9
        
        # 0A. Vertical Means [M Vert]
        # -----------------------------
        # Removing " V " (keeping alignment by spaces)
        v_row_txt = [indent_str, " " * FAC_WIDTH, left_padding_str, "   "] 
        v_row_tsv = ["M Vert Right", ""] + [""] * 7 + [""] # Empty for V
        
        
        for pos in range(1, 5):
            # Size Mean (Column)
            keys = [f'right_{pos}_totright_{tot}' for tot in range(pos, 5)]
            gm = calc_gm(keys)
            txt = f"{gm:.3f}" if gm is not None else "N/A"
            v_row_txt.append(txt.center(VAL_WIDTH))
            v_row_tsv.append(f"{gm:.3f}" if gm is not None else "")
            
            # Factors (Vertical Row: ONLY Horizontal Factors averaged)
            if pos < 4 and (show_horizontal_factors or show_diagonal_factors):
                h_keys = []
                for tot in range(pos + 1, 5):
                    h_keys.append(f'factor_right_{pos+1}_totright_{tot}_vs_right_{pos}_totright_{tot}')
                
                h_gm = calc_gm(h_keys)
                
                parts = []
                tsv_parts = []
                if h_gm is not None:
                    is_outward = get_growth_direction('right', arrow_direction) == 'outward'
                    val = h_gm if is_outward else (1.0 / h_gm if h_gm != 0 else 1.0)
                    sym = get_arrow('right', 'horizontal', arrow_direction)
                    s = f"×{val:.2f}{sym}"
                    parts.append(s)
                    tsv_parts.append(s)
                
                final_str = " ".join(parts)
                v_row_txt.append(final_str.center(FAC_WIDTH) if parts else E_FAC)
                v_row_tsv.append(" ".join(tsv_parts) if tsv_parts else "")

        # Append Extra Empty Column for M Vert
        v_row_txt.append(E_FAC) # Placeholder for Extra Col
        v_row_tsv.append("")

        lines.append("".join(v_row_txt) + "  [M Vert]")
        tsv_rows.append(v_row_tsv)
        
        # 0B. Diagonal Means [M Diag]
        # -----------------------------
        d_row_txt = [indent_str, " " * FAC_WIDTH, left_padding_str, "   "] # Remove V
        d_row_tsv = ["M Diag Right", ""] + [""] * 7 + [""] # Remove V
        
        # Calculate Specific Diagonal Means
        
        # 1. Longest Diagonal (Len 3): R1T1..R4T4 (Diag 0)
        # Factors: R1->R2 (T2), R2->R3 (T3), R3->R4 (T4)
        # key: factor_right_{u}_tot_{t}_vs_right_{u-1}_tot_{t-1}
        # u=2..4. t=u.
        keys_fac_d3 = []
        keys_sz_d3 = [f'right_1_totright_1', f'right_2_totright_2', f'right_3_totright_3', f'right_4_totright_4']
        for u in range(2, 5):
             keys_fac_d3.append(f'factor_right_{u}_totright_{u}_vs_right_{u-1}_totright_{u-1}')
        gm_fac_d3 = calc_gm(keys_fac_d3)
        gm_sz_d3 = calc_gm(keys_sz_d3)
        
        # 2. Medium Diagonal (Len 2): R1T2..R3T4 (Diag 1)
        # Factors: R1->R2 (T3), R2->R3 (T4)
        # u=2..3. t=u+1.
        keys_fac_d2 = []
        keys_sz_d2 = [f'right_1_totright_2', f'right_2_totright_3', f'right_3_totright_4']
        for u in range(2, 4):
             keys_fac_d2.append(f'factor_right_{u}_totright_{u+1}_vs_right_{u-1}_totright_{u}')
        gm_fac_d2 = calc_gm(keys_fac_d2)
        gm_sz_d2 = calc_gm(keys_sz_d2)
        
        # 3. Shortest Diagonal (Len 1): R1T3..R2T4 (Diag 2)
        # Factors: R1->R2 (T4)
        # u=2. t=u+2.
        keys_fac_d1 = []
        keys_sz_d1 = [f'right_1_totright_3', f'right_2_totright_4']
        keys_fac_d1.append(f'factor_right_{2}_totright_{4}_vs_right_{1}_totright_{3}')
        gm_fac_d1 = calc_gm(keys_fac_d1)
        gm_sz_d1 = calc_gm(keys_sz_d1)
        
        # Build Row Cells explicitly (8 columns)
        for i in range(8):
            if i == 3:
                fac, sz = gm_fac_d1, gm_sz_d1
            elif i == 5:
                fac, sz = gm_fac_d2, gm_sz_d2
            elif i == 7:
                fac, sz = gm_fac_d3, gm_sz_d3
            else:
                fac, sz = None, None
            
            if fac is not None:
                is_outward = get_growth_direction('right', arrow_direction) == 'outward'
                val = fac if is_outward else (1.0 / fac if fac != 0 else 1.0)
                sym = get_arrow('right', 'diagonal', arrow_direction)
                s = f"{sz:.2f} ×{val:.2f}{sym}" if sz is not None else f"×{val:.2f}{sym}"
                d_row_txt.append(s.center(FAC_WIDTH if i % 2 == 1 or i == 7 else VAL_WIDTH))
                d_row_tsv.append(s)
            else:
                d_row_txt.append(E_FAC if i % 2 == 1 or i == 7 else E_VAL)
                d_row_tsv.append("")

        lines.append("".join(d_row_txt) + "  [M Diag]")
        tsv_rows.append(d_row_tsv)
        lines.append("-" * 120)

    # ---------------------------------------------------------
    # 1. Right Dependents
    # ---------------------------------------------------------
    for tot in [4, 3, 2, 1]:
        # 1a. Print Current Row Values
        parts = []
        parts.append(" V") 
        
        # TSV Data for Right side
        tsv_right = ["V"]
        
        prev_val = None
        
        for pos in range(1, tot + 1):
            key = f'right_{pos}_totright_{tot}'
            val = position_averages.get(key, float('nan'))
            
            # Format Value
            if isinstance(val, float) and val != val:
                val_str = "N/A".center(VAL_WIDTH)
                tsv_val = "N/A"
                val = None
            else:
                val_str = f"{val:>{VAL_WIDTH}.3f}" if val is not None else "N/A".center(VAL_WIDTH)
                tsv_val = f"{val:.3f}" if val is not None else "N/A"
            
            # Horizontal Factor (between prev and current)
            if pos > 1 and show_horizontal_factors:
                fac_parts = []
                tsv_fac_parts = []
                
                # 1. Calculate Factor (Arrow)
                factor_val = None
                factor_str = ""
                
                # Check arrow direction logic first to get factor
                key_b = f'right_{pos}_totright_{tot}'
                key_a = f'right_{pos-1}_totright_{tot}'
                fac_key = f'factor_{key_b}_vs_{key_a}'
                geo_factor = position_averages.get(fac_key)
                if geo_factor is not None:
                    factor_val = geo_factor
                elif prev_val is not None and val is not None and prev_val != 0:
                    factor_val = val / prev_val
                          
                if factor_val is not None:
                    is_outward = get_growth_direction('right', arrow_direction) == 'outward'
                    val_factor_display = factor_val if is_outward else (1.0 / factor_val if factor_val != 0 else 1.0)
                    sym = get_arrow('right', 'horizontal', arrow_direction)
                    factor_str = f"×{val_factor_display:.2f}{sym}"
                          
                if factor_str:
                    fac_parts.append(factor_str)
                    tsv_fac_parts.append(factor_str) # Use the string with symbols

                # 2. Ordering Triples logic
                if show_ordering_triples and ordering_stats:
                    pair_idx = pos - 2
                    o_key = ('right', tot, pair_idx)
                    o_data = ordering_stats.get(o_key)
                    
                    if o_data:
                        total = o_data['lt'] + o_data['eq'] + o_data['gt']
                        if total > 0:
                            lt = o_data['lt'] / total * 100
                            eq = o_data['eq'] / total * 100
                            gt = o_data['gt'] / total * 100
                            # Format: (<66=14>21)
                            trip_str = f"(<{lt:.0f}={eq:.0f}>{gt:.0f})"
                            fac_parts.append(trip_str)
                            tsv_fac_parts.append(f"(<{lt:.1f}={eq:.1f}>{gt:.1f})")
                
                # Combine
                if fac_parts:
                    final_fac_str = " ".join(fac_parts)
                    # Adjust width? Standard FAC_WIDTH might be too small (14).
                    # If both are present: "x1.32 -> 12<5=83>" is approx 18 chars.
                    # We might need to center it in a wider slot OR just let it take space?
                    # The function relies on fixed width formatting (center(FAC_WIDTH)).
                    # If text is longer than width, center() does nothing.
                    # This might break alignment if we don't increase FAC_WIDTH or handle total string construction differently.
                    # But Python formatting usually handles overflow by just pushing.
                    pass
                else:
                    final_fac_str = E_FAC.strip() # maintain emptiness if nothing
                
                # Update cells
                # If we output a very long string, it might shift dependent columns visually if we rely on simple concat.
                # However, cells are just joined: "".join(parts).
                # So alignment of diagonals (which are on next line) depends on FAC_WIDTH matching VAL_WIDTH etc.
                # If this row expands, diagonals won't match up unless we change global constants.
                # For now, let's just center it.
                parts.append(final_fac_str.center(max(FAC_WIDTH, len(final_fac_str))))
                tsv_right.append(" ".join(tsv_fac_parts))
            
            parts.append(val_str)
            tsv_right.append(tsv_val)
            prev_val = val
            
        right_str = "".join(parts)
        
        # Add disorder percentage if provided (legacy)
        suffix_str = ""
        if disorder_percentages is not None:
            disorder_pct = disorder_percentages.get(('right', tot))
            if disorder_pct is not None:
                suffix_str += f"  [unordered {disorder_pct:.1f}%]"
                
        # Add Row Average
        row_avg = None # Initialize for scope
        if show_row_averages:
            avg_key = f'average_totright_{tot}'
            row_avg = position_averages.get(avg_key)
            if row_avg is not None:
                suffix_str += f"  [GM: {row_avg:.3f}]"
        
        # Add N if available
        n_count = None # Initialize for scope
        # Add N if available
        n_count = None # Initialize for scope
        if ordering_stats:
             t_key = ('right', tot, 'total')
             n_count = ordering_stats.get(t_key)
             
             if n_count is None and tot >= 2:
                 o_key = ('right', tot, 0)
                 o_data = ordering_stats.get(o_key)
                 if o_data:
                     n_count = o_data['lt'] + o_data['eq'] + o_data['gt']
        
        if n_count is not None:
             suffix_str += f" [N={n_count}]"
        
        # Add Slope
        slope_val = None # Initialize for scope
        if tot >= 2:
             y_vals = []
             valid = True
             for pos in range(1, tot + 1):
                 v = position_averages.get(f'right_{pos}_totright_{tot}')
                 if v is None: valid=False; break
                 y_vals.append(v)
             if valid:
                 try:
                     slope_val, _ = np.polyfit(np.arange(tot), y_vals, 1)
                     suffix_str += f" [Slope: {slope_val:+.2f}]"
                 except: pass

        # Pad right_str to fixed width
        # Width: ' V' (2) + HALF_WIDTH (content)
        target_width = 2 + HALF_WIDTH
        right_str = right_str.ljust(target_width)
        
        lines.append(f"R tot={tot}: {' ' * FAC_WIDTH}{LEFT_PAD}{right_str}{suffix_str}")
        
        # Build TSV Row
        # Left side empty for R rows
        row_label = f"R tot={tot}"
        # Fixed 7 slots for Left: L4, F, L3, F, L2, F, L1.
        left_empty = [""] * 7
        
        # Right side can be variable length. Pad to 7 (R1..R4)
        # Standard: R1, F, R2, F, R3, F, R4. (7 slots)
        # tsv_right currently has: V, R1, [F, R2], [F, R3]...
        # Let's fix loop logic to standardize
        # It pushed V, then (if pos>1 F), Val.
        # Order: V, R1, F12, R2, F23, R3, F34, R4.
        # tsv_right content: "V", "R1", "F12", "R2"...
        # We need to pad tsv_right to max length 8 (V + 7).
        while len(tsv_right) < 8: # Pad R columns 
            tsv_right.append("")
            
        if show_row_averages:
            stats_parts = []
            if row_avg is not None:
                stats_parts.append(f"[GM: {row_avg:.3f}]")
            if n_count is not None:
                stats_parts.append(f"[N={n_count}]")
            if slope_val is not None:
                 stats_parts.append(f"[Slope: {slope_val:+.2f}]")
            
            # Shift Stats Right (Extra Empty Column)
            full_tsv_row = [row_label, ""] + left_empty + tsv_right
            full_tsv_row.append("")
            full_tsv_row.append(" ".join(stats_parts))
        else:
            full_tsv_row = [row_label, ""] + left_empty + tsv_right
        
        tsv_rows.append(full_tsv_row)
        
        # 1b. Print Diagonal Factors (Connection to Next Row: tot-1)
        if tot > 1 and show_diagonal_factors:
            diag_parts = []
            diag_parts.append("   ") 
            
            # TSV Diag Row
            tsv_diag_right = [""] # Under V
            
            for pos in range(1, tot): 
                diag_parts.append(E_VAL) 
                tsv_diag_right.append("") # Under R_pos
                
                if show_horizontal_factors:
                    src_key = f'right_{pos}_totright_{tot-1}'
                    tgt_key = f'right_{pos+1}_totright_{tot}'
                    src_val = position_averages.get(src_key, None)
                    tgt_val = position_averages.get(tgt_key, None)
                    
                    if src_val is not None and tgt_val is not None and src_val != 0:
                        factor = None
                        fac_key = f'factor_{tgt_key}_vs_{src_key}'
                        geo_factor = position_averages.get(fac_key)
                        
                        if geo_factor is not None:
                            factor_val = geo_factor if is_outward else (1.0 / geo_factor if geo_factor != 0 else 1.0)
                        else:
                            factor_val = tgt_val / src_val if is_outward else src_val / tgt_val
                        
                        sym = get_arrow('right', 'diagonal', arrow_direction)
                        diag_str = f"×{factor_val:.2f} {sym}".center(FAC_WIDTH)
                        tsv_fac = f"×{factor_val:.2f} {sym}"
                    else:
                        diag_str = E_FAC
                        tsv_fac = ""
                    diag_parts.append(diag_str)
                    tsv_diag_right.append(tsv_fac)
                
            diag_parts.append(E_VAL) 
            while len(tsv_diag_right) < 8: tsv_diag_right.append("")
                
            lines.append(f"        {' ' * FAC_WIDTH}{LEFT_PAD}{''.join(diag_parts)}")
            
            tsv_rows.append([f"Diag R{tot}-{tot-1}", ""] + left_empty + tsv_diag_right)

    lines.append("-" * 120)
    tsv_rows.append([""])

    # ---------------------------------------------------------
    # 1.5. XVX Configuration
    # ---------------------------------------------------------
    l_xvx = position_averages.get('xvx_left_1')
    r_xvx = position_averages.get('xvx_right_1')
    
    if l_xvx is not None or r_xvx is not None:
         # Text output
         # Layout: "X V X :   ...  L1  V  R1  ..."
         # Left Grid: L1 is at inner most slot (fac=6 if factors, 3 if not)
         lines.append("-" * 120)
         
         # Left Grid Cells
         l_cells = [E_VAL for _ in range(4)]
         tsv_left = [""] * 7
         
         l_idx = 3 # L1
         l_cells[l_idx] = f"{l_xvx:.3f}".center(VAL_WIDTH) if l_xvx else E_VAL
         
         if show_horizontal_factors:
             # L4 F L3 F L2 F L1
             # val at index 6.
             l_cells = [E_VAL, E_FAC, E_VAL, E_FAC, E_VAL, E_FAC, E_VAL]
             l_cells[6] = f"{l_xvx:.3f}".center(VAL_WIDTH) if l_xvx else E_VAL
             tsv_left[6] = f"{l_xvx:.3f}" if l_xvx else ""
         else:
             tsv_left = [""] * 4
             tsv_left[3] = f"{l_xvx:.3f}" if l_xvx else ""
             
         left_block = "".join(l_cells)
         # Pad for missing factors if needed? Handled by initial join.
         
         # Right Grid
         r_cells = [f"{r_xvx:.3f}".center(VAL_WIDTH) if r_xvx else E_VAL]
         tsv_right = ["V", f"{r_xvx:.3f}" if r_xvx else ""]
         
         if show_horizontal_factors:
              # R1 F R2 F R3 F R4 Extra
              r_cells.append(E_FAC) # F12
              
              # Factor L1 -> R1
              if l_xvx and r_xvx and l_xvx != 0:
                  fac = r_xvx / l_xvx
                  fac_key = 'factor_xvx_right_1_vs_xvx_left_1'
                  geo_factor = position_averages.get(fac_key)
                  if geo_factor: fac = geo_factor
                  else: fac = r_xvx / l_xvx
                  
                  # Apply inward logic
                  is_outward = get_growth_direction('right', arrow_direction) == 'outward'
                  val = fac if is_outward else (1.0 / fac if fac != 0 else 1.0)
                  sym = get_arrow('right', 'horizontal', arrow_direction)
                  f_str = f"×{val:.2f}{sym}"
                  r_cells[1] = f_str.center(FAC_WIDTH)
                  tsv_right.append(f_str)
              else:
                  tsv_right.append("")
                  
              while len(r_cells) < 8: r_cells.append(E_FAC if len(r_cells) % 2 == 1 else E_VAL)
              while len(tsv_right) < 8: tsv_right.append("")
         
         # Add N if available
         n_str = ""
         n_count = None
         if ordering_stats:
              n_key = ('xvx', 2, 'total')
              n_count = ordering_stats.get(n_key)
              if n_count: n_str = f" [N={n_count}]"
         
         lines.append(f"X V X :   {' ' * FAC_WIDTH}{left_block} V { ''.join(r_cells)}{n_str}")
         tsv_rows.append(["X V X", ""] + tsv_left + tsv_right + [""])
         lines.append("-" * 120)
         tsv_rows.append([""])

    # ---------------------------------------------------------
    # 2. Left Dependents
    # ---------------------------------------------------------
    for tot in [1, 2, 3, 4]:
        # 2a. Print Current Row Values
        # TSV Left Row
        tsv_left = [""] * 7
        
        row_cells_text = [] # For text output
        cells = [""] * (4 + 3) if show_horizontal_factors else [""] * 4
        # Fill with spaces
        for i in range(len(cells)):
             cells[i] = E_FAC if (show_horizontal_factors and i % 2 == 1) else E_VAL

        prev_val_inner = None 
        
        for pos in range(1, tot + 1):
             key = f'left_{pos}_totleft_{tot}'
             val = position_averages.get(key, None)
             
             if show_horizontal_factors:
                 val_idx = 6 - (pos - 1) * 2
             else:
                 val_idx = 3 - (pos - 1)
             
             if val is not None:
                 cells[val_idx] = f"{val:>{VAL_WIDTH}.3f}"
                 tsv_left[val_idx] = f"{val:.3f}"
             else:
                 cells[val_idx] = "N/A".center(VAL_WIDTH)
                 val = None

             if pos > 1 and show_horizontal_factors:
                 fac_idx = val_idx + 1 # Slot to the right
                 
                 fac_parts = []
                 tsv_fac_parts = []
                 # 1. Calculate Factor (Arrow)
                 factor_val = None
                 factor_str = ""
                 
                 key_b = f'left_{pos}_totleft_{tot}'
                 key_a = f'left_{pos-1}_totleft_{tot}'
                 fac_key = f'factor_{key_b}_vs_{key_a}'
                 geo_factor = position_averages.get(fac_key)
                 if geo_factor is not None:
                    factor_val = geo_factor
                 elif prev_val_inner is not None and val is not None and prev_val_inner != 0:
                    factor_val = val / prev_val_inner
                          
                 if factor_val is not None:
                    is_outward = get_growth_direction('left', arrow_direction) == 'outward'
                    val_factor_display = factor_val if is_outward else (1.0 / factor_val if factor_val != 0 else 1.0)
                    sym = get_arrow('left', 'horizontal', arrow_direction)
                    factor_str = f"×{val_factor_display:.2f}{sym}"
                          
                 if factor_str:
                    fac_parts.append(factor_str)
                    tsv_fac_parts.append(factor_str) # Use the string with symbols

                 # 2. Ordering Triples logic
                 if show_ordering_triples and ordering_stats:
                     # Fix: For Left side, neighbors in "L(pos) -> L(pos-1)" visual order
                     # correspond to list indices (tot-pos) and (tot-pos+1).
                     # pair_idx in ordering_stats matches the first element's index in the sorted list.
                     # List is [L_k ... L_1] (visual left-to-right).
                     # At pos (displaying L_pos, factor between L_pos and L_{pos-1}), 
                     # we want the pair (L_pos, L_{pos-1}).
                     # L_pos corresponds to index (tot - pos).
                     pair_idx = tot - pos
                     o_key = ('left', tot, pair_idx)
                     o_data = ordering_stats.get(o_key)
                     
                     if o_data:
                        total = o_data['lt'] + o_data['eq'] + o_data['gt']
                        if total > 0:
                            lt = o_data['lt'] / total * 100
                            eq = o_data['eq'] / total * 100
                            gt = o_data['gt'] / total * 100
                            # Format: (<12=5>83)
                            trip_str = f"(<{lt:.0f}={eq:.0f}>{gt:.0f})"
                            fac_parts.append(trip_str)
                            tsv_fac_parts.append(f"(<{lt:.1f}={eq:.1f}>{gt:.1f})")
                            
                 # Combine
                 if fac_parts:
                     final_fac_str = " ".join(fac_parts)
                 else:
                     final_fac_str = E_FAC.strip()

                 cells[fac_idx] = final_fac_str.center(max(FAC_WIDTH, len(final_fac_str)))
                 tsv_left[fac_idx] = " ".join(tsv_fac_parts)

             prev_val_inner = val
             
        left_str = "".join(cells)
        
        # Add disorder percentage if provided (legacy)
        suffix_str = ""
        if disorder_percentages is not None:
            disorder_pct = disorder_percentages.get(('left', tot))
            if disorder_pct is not None:
                suffix_str += f"  [unordered {disorder_pct:.1f}%]"
                
        # Add Row Average
        row_avg = None # Initialize for scope
        if show_row_averages:
            avg_key = f'average_totleft_{tot}'
            row_avg = position_averages.get(avg_key)
            if row_avg is not None:
                suffix_str += f"  [GM: {row_avg:.3f}]"
        
        # Add N
        n_count = None # Initialize for scope
        # Add N
        n_count = None # Initialize for scope
        if ordering_stats:
             t_key = ('left', tot, 'total')
             n_count = ordering_stats.get(t_key)
             
             if n_count is None and tot >= 2:
                 o_key = ('left', tot, 0)
                 o_data = ordering_stats.get(o_key)
                 if o_data:
                     n_count = o_data['lt'] + o_data['eq'] + o_data['gt']
                     
        if n_count is not None:
             suffix_str += f" [N={n_count}]"
                
        # Add Slope
        slope_val = None # Initialize for scope
        if tot >= 2:
             y_vals = []
             valid = True
             # Visual order: tot down to 1
             for pos in range(tot, 0, -1):
                 v = position_averages.get(f'left_{pos}_totleft_{tot}')
                 if v is None: valid=False; break
                 y_vals.append(v)
             if valid:
                 try:
                     slope_val, _ = np.polyfit(np.arange(tot), y_vals, 1)
                     suffix_str += f" [Slope: {slope_val:+.2f}]"
                 except: pass

        # Pad Stats Right (Extra Empty Column)
        right_pad = " " * HALF_WIDTH
        extra_gap = " " * FAC_WIDTH 
        lines.append(f"L tot={tot}: {' ' * FAC_WIDTH}{left_str} V{right_pad}{extra_gap}{suffix_str}")
        
        tsv_row = [f"L tot={tot}", ""] + tsv_left + ["V"] + [""]*7
        
        # Append Stats to TSV
        # Shift Stats Right (Extra Empty Column)
        tsv_row.append("")
        
        if show_row_averages:
            stats_parts = []
            if row_avg is not None:
                stats_parts.append(f"[GM: {row_avg:.3f}]")
            if n_count is not None:
                stats_parts.append(f"[N={n_count}]")
            if slope_val is not None:
                stats_parts.append(f"[Slope: {slope_val:+.2f}]")
            
            if stats_parts:
                tsv_row.append(" ".join(stats_parts))
            else:
                tsv_row.append("") # Empty stats if shown but none found
        
        tsv_rows.append(tsv_row)
        
        # 2b. Print Diagonal Factors
        if tot < 4 and show_diagonal_factors and show_horizontal_factors:
            diag_cells = [""] * 7
            tsv_diag_left = [""] * 7
            for i in range(7): diag_cells[i] = E_FAC if i % 2 == 1 else E_VAL
            
            for pos in range(1, tot + 1):
                tgt_key = f'left_{pos}_totleft_{tot}'
                src_key = f'left_{pos+1}_totleft_{tot+1}'
                tgt_val = position_averages.get(tgt_key, None)
                src_val = position_averages.get(src_key, None)
                
                if pos + 1 > 4: continue
                fac_idx = (6 - (pos - 1) * 2) - 1 # Slot left of pos
                
                if src_val is not None and tgt_val is not None and src_val != 0:
                     factor = None
                     fac_key = f'factor_{tgt_key}_vs_{src_key}'
                     geo_factor = position_averages.get(fac_key)
                     
                     if geo_factor is not None:
                          val_in = geo_factor
                     else:
                          val_in = tgt_val / src_val 

                     if arrow_direction == 'left_to_right':
                          factor_val = val_in
                     else:
                          is_outward = get_growth_direction('left', arrow_direction) == 'outward'
                          factor_val = val_in if is_outward else (1.0 / val_in if val_in != 0 else 1.0)
                          
                     # Arrow Symbol
                     sym = get_arrow('left', 'diagonal', arrow_direction)
                     diag_cells[fac_idx] = f"×{factor_val:.2f} {sym}".center(FAC_WIDTH)
                     tsv_diag_left[fac_idx] = f"{factor_val:.2f} {sym}"
                
            lines.append(f"                       {' ' * FAC_WIDTH}{''.join(diag_cells)}   ")
            tsv_rows.append([f"Diag L{tot}-{tot+1}", ""] + tsv_diag_left + [""] * 9)

    # ---------------------------------------------------------
    # 3. Left Marginal Means (Bottom)
    if show_marginal_means:
        calc_gm = calc_marginal_gm_txt # alias
        lines.append("-" * 120)
        
        # ---------------------------------------------------------
        # 3. Left Marginal Means (Bottom)
        # ---------------------------------------------------------
        l_vert_size = {}
        l_vert_fac = {}
        
        for pos in range(1, 5):
            keys = [f'left_{pos}_totleft_{tot}' for tot in range(pos, 5)]
            l_vert_size[pos] = calc_gm(keys)
            
            if pos > 1:
                h_keys = [f'factor_left_{pos}_totleft_{tot}_vs_left_{pos-1}_totleft_{tot}' for tot in range(pos, 5)]
                d_keys = [f'factor_left_{pos-1}_totleft_{tot}_vs_left_{pos}_totleft_{tot+1}' for tot in range(pos-1, 4)]
                
                h_gm = calc_gm(h_keys)
                d_gm = calc_gm(d_keys)
                
                vals = []
                if h_gm is not None:
                    is_outward = get_growth_direction('left', arrow_direction) == 'outward'
                    v = h_gm if is_outward else (1.0 / h_gm if h_gm != 0 else 1.0)
                    vals.append(f"×{v:.2f}{get_arrow('left', 'horizontal', arrow_direction)}")
                if d_gm is not None:
                    # Stored d_gm is Inner/Outer (L_{P-1}/L_P).
                    # We want Growth (Outer/Inner). So we need 1/d_gm.
                    # If left_to_right: L(Outer) -> L(Inner) is inward.
                    # Normal Inward logic: 1/val.
                    # So (1/d_gm) is Growth.
                    # If we just use standard inversion logic:
                    # is_outward = False. val = 1.0 / val_in.
                    # If val_in = d_gm (Inner/Outer). Then val = Outer/Inner. Correct.
                    is_outward = get_growth_direction('left', arrow_direction) == 'outward'
                    val = d_gm if is_outward else (1.0 / d_gm if d_gm != 0 else 1.0)
                    
                    # Force fix for left_to_right if needed?
                    # logic: d_gm is L_{P-1}/L_P (Inner/Outer).
                    # left_to_right: we want L_P -> L_{P-1}. (Growth).
                    # ie Outer -> Inner.
                    # Growth = Inner/Outer? No. Outer=Small, Inner=Large?
                    # English: L4=1.1, L1=1.2. Inner > Outer.
                    # So Inner/Outer > 1.
                    # d_gm is Inner/Outer. So d_gm > 1.
                    # If d_gm > 1, and we want Growth, just display d_gm?
                    # get_growth_direction('left', 'left_to_right') => 'inward'.
                    # is_outward = False.
                    # val = 1.0 / d_gm. => < 1. WRONG.
                    # If d_gm is ALREADY relative to flow?
                    # Let's trust standard d_gm key: factor_L_{P-1}_vs_L_{P}. (Target vs Source).
                    # If Target=Inner, Source=Outer. Then d_gm = Inner/Outer.
                    # If Inward flow (Outer->Inner), then Ratio = End/Start = Inner/Outer.
                    # So d_gm IS the Growth factor.
                    # But standard logic inverts if "inward".
                    # So we should PRE-INVERT d_gm so that standard logic flips it back?
                    # OR just override standard logic.
                    if arrow_direction == 'left_to_right':
                         val = d_gm
                    else:
                         val = d_gm if is_outward else (1.0 / d_gm if d_gm != 0 else 1.0)

                    vals.append(f"×{val:.2f}{get_arrow('left', 'diagonal', arrow_direction)}")
                l_vert_fac[pos] = " ".join(vals)

        # 3A. Diagonal Means [M Diag Left]
        raw_d_size = {}
        for k in range(4):
            keys = [f'left_{pos}_totleft_{pos+k}' for pos in range(1, 4-k+1)]
            raw_d_size[k] = calc_gm(keys)
            
        # Keys are factor_Inner_vs_Outer.
        # d_gm values from calc_gm are Inner/Outer.
        # If left_to_right (Inward flow), Growth = Inner/Outer.
        # So we want raw GM.
        def get_inv_gm_smart(key_list):
            val = calc_gm(key_list)
            if not val: return None
            # If left_to_right, return raw val (Inner/Outer).
            if arrow_direction == 'left_to_right': return val
            # Else invert?
            return 1.0 / val
            
        gm_fac_d3 = get_inv_gm_smart([f'factor_left_{u-1}_totleft_{u-1}_vs_left_{u}_totleft_{u}' for u in range(2, 5)])
        gm_sz_d3 = raw_d_size.get(0)
        gm_fac_d2 = get_inv_gm_smart([f'factor_left_{u-1}_totleft_{u}_vs_left_{u}_totleft_{u+1}' for u in range(2, 4)])
        gm_sz_d2 = raw_d_size.get(1)
        gm_fac_d1 = get_inv_gm_smart([f'factor_left_{1}_totleft_{3}_vs_left_{2}_totleft_{4}'])
        gm_sz_d1 = raw_d_size.get(2)

        d_row_txt = [indent_str]
        d_row_tsv = ["M Diag Left"]
        
        if gm_fac_d3:
            # Already handled logic in get_inv_gm_smart
            # Just display v directly if left_to_right?
            # Or pass through standard logic?
            # if left_to_right, gm_fac_d3 is already Growth.
            # But standard logic below (is_outward...) might flip it again.
            if arrow_direction == 'left_to_right':
                v = gm_fac_d3
            else:
                 is_outward = get_growth_direction('left', arrow_direction) == 'outward'
                 v = gm_fac_d3 if is_outward else (1.0 / gm_fac_d3 if gm_fac_d3 != 0 else 1.0)
                 
            sym = get_arrow('left', 'diagonal', arrow_direction)
            m_diag_str = f"{gm_sz_d3:.2f} ×{v:.2f}{sym}"
        else:
            m_diag_str = ""
        d_row_txt.append(m_diag_str.center(FAC_WIDTH) if m_diag_str else " " * FAC_WIDTH)
        d_row_tsv.append(m_diag_str)
        
        # Grid loop for 7 cells
        l_d_grid_parts = []
        l_d_grid_tsv = []
        for i in range(4, 0, -1):
            l_d_grid_parts.append(E_VAL) 
            l_d_grid_tsv.append("")
            
            if i > 1:
                fac, sz = None, None
                if i == 4: fac, sz = gm_fac_d2, gm_sz_d2
                elif i == 3: fac, sz = gm_fac_d1, gm_sz_d1
                
                if fac:
                    if arrow_direction == 'left_to_right':
                         v = fac
                    else:
                         is_outward = get_growth_direction('left', arrow_direction) == 'outward'
                         v = fac if is_outward else (1.0 / fac if fac != 0 else 1.0)
                         
                    sym = get_arrow('left', 'diagonal', arrow_direction)
                    s = f"{sz:.2f} ×{v:.2f}{sym}"
                    l_d_grid_parts.append(s.center(FAC_WIDTH))
                    l_d_grid_tsv.append(s)
                else:
                    l_d_grid_parts.append(E_FAC)
                    l_d_grid_tsv.append("")
        
        d_row_txt.extend(l_d_grid_parts)
        d_row_txt.append(" V ")
        d_row_tsv.extend(l_d_grid_tsv)
        d_row_tsv.append("V")
        
        # Padding for Right Side
        d_row_txt.append(" " * (HALF_WIDTH + FAC_WIDTH))
        d_row_tsv.extend([""] * 8)
        
        lines.append("".join(d_row_txt) + "  [M Diag Left]")
        tsv_rows.append(d_row_tsv)


        # 3B. Vertical Means [M Vert Left]
        v_row_txt = [indent_str, " " * FAC_WIDTH] 
        v_row_tsv = ["M Vert Left", ""] 
        
        for i in range(4, 0, -1):
            gm = l_vert_size.get(i)
            v_row_txt.append(f"{gm:.3f}".center(VAL_WIDTH) if gm is not None else E_VAL)
            v_row_tsv.append(f"{gm:.3f}" if gm is not None else "")
            
            if i > 1:
                fac_str = l_vert_fac.get(i, "")
                v_row_txt.append(fac_str.center(FAC_WIDTH) if fac_str else E_FAC)
                v_row_tsv.append(fac_str)
        
        v_row_txt.append(" V ")
        v_row_tsv.append("V")
        # Padding for Right Side: HALF_WIDTH content + FAC_WIDTH Extra
        v_row_txt.append(" " * (HALF_WIDTH + FAC_WIDTH))
        v_row_tsv.extend([""] * 8)
        
        lines.append("".join(v_row_txt) + "  [M Vert Left]")
        tsv_rows.append(v_row_tsv)



    lines.append("=" * 120)
    lines.append(f"Mode: Factors={show_horizontal_factors}, Diagonals={show_diagonal_factors}, Direction={arrow_direction}")
    lines.append("========================================================================================================================")
    
    # Save TSV
    if save_tsv and output_dir:
        # Auto-generate filename if not provided
        if filename is None:
            parts = ['verb_centered_table']
            if show_horizontal_factors:
                parts.append('factors')
            if show_diagonal_factors:
                parts.append('diagonal')
            if show_horizontal_factors or show_diagonal_factors:
                parts.append(arrow_direction)
            filename = "_".join(parts) + ".tsv"
            
        try:
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            full_path = os.path.join(output_dir, filename)
            with open(full_path, 'w', encoding='utf-8') as f:
                for row in tsv_rows:
                    f.write("\t".join(row) + "\n")
            lines.append(f"Table saved to: {full_path}")
        except Exception as e:
            lines.append(f"Error saving TSV: {e}")
            
    return "\n".join(lines)


def _save_excel_verb_centered_table_legacy(grid_rows, output_path):
    """
    LEGACY IMPLEMENTATION - Saves the grid to an Excel file with conditional formatting.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Color
        from openpyxl.utils import get_column_letter
        from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont
    except ImportError:
        print("Error: openpyxl not installed. Cannot save Excel.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Verb Centered Analysis"

    # Styles
    bold_font = Font(bold=True)
    grey_font = Font(color="555555") 
    red_font = Font(color="FF0000", bold=True)
    
    # alignments
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    for r_idx, row in enumerate(grid_rows, 1):
        for c_idx, cell in enumerate(row, 1):
            c = ws.cell(row=r_idx, column=c_idx)
            
            # 1. Content
            if cell.rich_text:
                # Construct CellRichText
                rt = CellRichText()
                for (text, color_hex, is_bold) in cell.rich_text:
                    # Convert to InlineFont
                    if_font = InlineFont(color=color_hex, b=is_bold)
                    rt.append(TextBlock(font=if_font, text=text))
                c.value = rt
            elif cell.value is not None:
                c.value = cell.value
                if isinstance(cell.value, (int, float)):
                    c.number_format = '0.000'
            else:
                c.value = cell.text
            
            # 2. Alignment
            if c_idx == 1:
                c.alignment = left_align
            else:
                c.alignment = center_align

            # 3. Base Font (if not RichText)
            # RichText overrides cell font for the parts, but base font might matter?
            # Actually RichText completely controls the content styling.
            if not cell.rich_text:
                if cell.cell_type == 'value':
                    c.font = bold_font
                elif cell.cell_type == 'factor':
                    # Fallback if logic failed or simple mode
                    if cell.is_astonishing: # Deprecated by RichText but keep for safety
                        c.font = red_font
                    else:
                        c.font = grey_font
                elif cell.cell_type == 'comment':
                    c.font = Font(italic=True, color="333333")
                elif cell.text == "V":
                    c.font = Font(bold=True, size=14)
            
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter 
        for cell in col:
            try:
                # Length of value str
                val_str = str(cell.value)
                if len(val_str) > max_length:
                    max_length = len(val_str)
            except:
                pass
        adjusted_width = (max_length + 2)
        # Cap width? Averages are small, factors are long.
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_path)
    # print(f"Saved Excel table to {output_path}")


def aggregate_ordering_stats(ordering_stats_dict):
    """
    Aggregates ordering statistics across multiple languages.
    """
    aggregated = {}
    for lang_stats in ordering_stats_dict.values():
        if not lang_stats: continue
        for key, counts in lang_stats.items():
            if isinstance(counts, int):
                # Aggregate totals
                if key not in aggregated:
                    aggregated[key] = 0
                aggregated[key] += counts
            else:
                if key not in aggregated:
                    aggregated[key] = {'lt': 0, 'eq': 0, 'gt': 0}
                aggregated[key]['lt'] += counts.get('lt', 0)
                aggregated[key]['eq'] += counts.get('eq', 0)
                aggregated[key]['gt'] += counts.get('gt', 0)
    return aggregated


def generate_mass_tables(all_langs_average_sizes, 
                         ordering_stats, 
                         metadata, 
                         vo_data=None, 
                         output_dir='data/tables',
                         arrow_direction='left_to_right',
                         extract_disorder_metrics=False):
    """
    Generates text/TSV tables and Excel files.
    
    Args:
        arrow_direction: Direction mode for growth factors. Options:
            - 'left_to_right': Left side inward (L4→L1→V), Right side outward (V→R1→R4)
            - 'right_to_left': Left side outward (V→L1→L4), Right side inward (R4→R1→V)
            - 'diverging' or 'outward': Both sides outward from V
            - 'inward' or 'converging': Both sides inward to V
        extract_disorder_metrics: If True, collect disorder extreme percentages for each language
        
    Returns:
        DataFrame with disorder metrics if extract_disorder_metrics=True, else None
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    lang_names = metadata.get('langNames', {})
    
    # Storage for disorder metrics if requested
    disorder_data = [] if extract_disorder_metrics else None
    
    # helper to process a subset of languages
    def process_subset(subset_langs, label, filename_label):
        if not subset_langs: return

        # Filter Data
        subset_avgs = {l: all_langs_average_sizes[l] for l in subset_langs if l in all_langs_average_sizes}
        subset_stats = {l: ordering_stats[l] for l in subset_langs if l in ordering_stats} if ordering_stats else None
        
        if not subset_avgs: return

        # Aggregate
        combined_avgs = compute_average_sizes_table(subset_avgs)
        combined_stats = aggregate_ordering_stats(subset_stats) if subset_stats else None
            
        # 1. Generate Standard TSV/Text
        filename_tsv = f"{filename_label}.tsv"
        format_verb_centered_table(
            combined_avgs,
            show_horizontal_factors=True,
            show_diagonal_factors=True,
            show_ordering_triples=True,
            show_row_averages=True,
            ordering_stats=combined_stats,
            arrow_direction=arrow_direction,
            save_tsv=True,
            output_dir=output_dir,
            filename=filename_tsv
        )
        
        # 2. Generate Excel with extracted grid
        grid = extract_verb_centered_grid(
            combined_avgs,
            show_horizontal_factors=True,
            show_diagonal_factors=True,
            arrow_direction=arrow_direction,
            ordering_stats=combined_stats,
            show_ordering_triples=True,
            show_row_averages=True
        )
        filename_xlsx = f"{filename_label}.xlsx"
        save_excel_verb_centered_table(grid, os.path.join(output_dir, filename_xlsx))

    print(f"Generating tables in {output_dir}...")
    
    all_langs = list(all_langs_average_sizes.keys())
    
    # Global
    process_subset(all_langs, "All Languages", "Table_All_Languages")
    
    # Individual
    for lang in all_langs:
        name = lang_names.get(lang, lang).replace(" ", "_").replace("/", "-")
        process_subset([lang], f"Language {name}", f"Helix_{name}_{lang}")
        
        # Extract disorder metrics if requested
        if extract_disorder_metrics and lang in ordering_stats:
            from verb_centered_computations import OrderingStatsFormatter, MarginalMeansCalculator
            formatter = OrderingStatsFormatter(ordering_stats[lang])
            
            # Get right side last pair aggregate (gt percentage)
            right_result = formatter.get_aggregate_right_last_pair_triple()
            right_extreme = right_result[2] if right_result else None  # gt percentage
            
            # Get left side first pair aggregate (gt percentage, after swap)
            left_result = formatter.get_aggregate_left_first_pair_triple()
            if left_result:
                lt_pct, eq_pct, gt_pct, total_n = left_result
                # Swap for left side (same as in display)
                left_extreme = lt_pct  # After swap, this is the extreme
            else:
                left_extreme = None
            
            # Extract extreme diagonal factors (M Diag Right rightmost, M Diag Left leftmost)
            marginals = MarginalMeansCalculator(all_langs_average_sizes[lang])
            
            # Right extreme diagonal factor (diag_idx 0 = R4 position, longest diagonal)
            diag_right = marginals.calc_diagonal_means_right()
            _, right_diag_factor = diag_right.get(0, (None, None))
            
            # Left extreme diagonal factor (diag_idx 0 = L4 position, longest diagonal)
            diag_left = marginals.calc_diagonal_means_left()
            _, left_diag_factor = diag_left.get(0, (None, None))
            
            disorder_data.append({
                'language_code': lang,
                'language_name': lang_names.get(lang, lang),
                'right_extreme_disorder': right_extreme,
                'left_extreme_disorder': left_extreme,
                'right_extreme_diag_factor': right_diag_factor,
                'left_extreme_diag_factor': left_diag_factor
            })
        
    # Families (IE / Non-IE)
    # Re-using previous logic but simpler iteration
    ie_langs = []
    non_ie_langs = []
    langnameGroup = metadata.get('langnameGroup', {})
    for lang in all_langs:
        name = lang_names.get(lang, lang)
        group = langnameGroup.get(name, 'Unknown')
        if group == 'Indo-European': ie_langs.append(lang)
        else: non_ie_langs.append(lang)
            
    process_subset(ie_langs, "Indo-European", "Table_Family_IndoEuropean")
    process_subset(non_ie_langs, "Non-Indo-European", "Table_Family_NonIndoEuropean")
    
    # Order
    if vo_data:
        vo_langs = []
        ov_langs = []
        ndo_langs = []
        for lang in all_langs:
            info = vo_data.get(lang)
            if info:
                vtype = info.get('vo_type')
                if not vtype and 'vo_score' in info:
                     score = info['vo_score']
                     if score is not None:
                         if score > 0.666: vtype = 'VO'
                         elif score < 0.333: vtype = 'OV'
                         else: vtype = 'NDO'
                if vtype == 'VO': vo_langs.append(lang)
                elif vtype == 'OV': ov_langs.append(lang)
                elif vtype == 'NDO': ndo_langs.append(lang)
        
        process_subset(vo_langs, "VO Languages", "Table_Order_VO")
        process_subset(ov_langs, "OV Languages", "Table_Order_OV")
        process_subset(ndo_langs, "NDO Languages", "Table_Order_NDO")
        
    print(f"Mass table generation complete. Output in: {output_dir}")
    
    # Return disorder metrics if collected
    if extract_disorder_metrics and disorder_data:
        import pandas as pd
        return pd.DataFrame(disorder_data)
    return None
