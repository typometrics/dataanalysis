"""
Formatters for verb-centered constituent size tables.

This module provides different output formatters (text, TSV, Excel)
for the TableStructure data model.
"""

import os
from typing import List
from verb_centered_model import TableStructure, CellData


class TextTableFormatter:
    """
    Formats TableStructure as fixed-width text.
    """
    
    def __init__(self, table: TableStructure):
        """
        Initialize formatter.
        
        Args:
            table: TableStructure to format
        """
        self.table = table
        self.VAL_WIDTH = 11
        self.FAC_WIDTH = 14
        self.LABEL_WIDTH = 12
    
    def format(self) -> str:
        """
        Generate text representation.
        
        Returns:
            Formatted text string
        """
        lines = []
        
        for row in self.table.rows:
            if not row:
                continue
            
            # Check if separator
            if len(row) == 1 and row[0].cell_type == 'separator':
                lines.append("-" * 120)
                continue
            
            # Format regular row
            line_parts = []
            for i, cell in enumerate(row):
                # Determine width
                if i == 0:  # Label column
                    width = self.LABEL_WIDTH
                elif cell.cell_type == 'factor':
                    width = self.FAC_WIDTH
                elif cell.cell_type == 'comment':
                    # Comments can be variable width
                    line_parts.append("  " + cell.text)
                    continue
                else:
                    width = self.VAL_WIDTH
                
                # Format cell
                text = cell.text if cell.text else ""
                
                # Handle V specially
                if text == "V" and i > 0:
                    formatted = text.center(width)
                elif i == 0:  # Labels left-aligned
                    formatted = text.ljust(width)
                else:  # Values and factors centered
                    formatted = text.center(width)
                
                line_parts.append(formatted)
            
            lines.append("".join(line_parts))
        
        return "\n".join(lines)


class TSVFormatter:
    """
    Formats TableStructure as tab-separated values.
    """
    
    def format(self, table: TableStructure) -> str:
        """
        Generate TSV representation.
        
        Args:
            table: TableStructure to format
            
        Returns:
            TSV string
        """
        rows = []
        
        for row in table.rows:
            if not row:
                continue
            
            # Convert separators to empty lines in TSV
            if len(row) == 1 and row[0].cell_type == 'separator':
                rows.append("")  # Add empty line for separator
                continue
            
            # Extract text from each cell
            row_texts = []
            for cell in row:
                text = cell.text if cell.text else ""
                row_texts.append(text)
            
            rows.append("\t".join(row_texts))
        
        return "\n".join(rows)
    
    def save(self, table: TableStructure, output_path: str):
        """
        Save TSV to file.
        
        Args:
            table: TableStructure to save
            output_path: Path to output file
        """
        content = self.format(table)
        
        # Ensure directory exists
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(content)


class ExcelFormatter:
    """
    Formats TableStructure as Excel with rich formatting.
    """
    
    def save(self, table: TableStructure, output_path: str):
        """
        Save as Excel file with formatting.
        
        Args:
            table: TableStructure to save
            output_path: Path to output .xlsx file
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont
        except ImportError:
            print("Error: openpyxl not installed. Cannot save Excel.")
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Verb Centered Analysis"
        
        # Styles
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        for r_idx, row in enumerate(table.rows, 1):
            if not row:
                continue
            
            # Skip separators in Excel
            if len(row) == 1 and row[0].cell_type == 'separator':
                continue
            
            for c_idx, cell_data in enumerate(row, 1):
                excel_cell = ws.cell(row=r_idx, column=c_idx)
                
                # Content
                if cell_data.rich_segments:
                    # Construct CellRichText
                    rt = CellRichText()
                    for (text, color_hex, is_bold) in cell_data.rich_segments:
                        if_font = InlineFont(color=color_hex, b=is_bold)
                        rt.append(TextBlock(font=if_font, text=text))
                    excel_cell.value = rt
                elif cell_data.value is not None:
                    excel_cell.value = cell_data.value
                    if isinstance(cell_data.value, (int, float)):
                        excel_cell.number_format = '0.000'
                else:
                    excel_cell.value = cell_data.text
                
                # Alignment
                if c_idx == 1:
                    excel_cell.alignment = left_align
                else:
                    excel_cell.alignment = center_align
                
                # Base font (if not rich text)
                if not cell_data.rich_segments:
                    if cell_data.cell_type in ['value', 'label']:
                        excel_cell.font = bold_font
                    elif cell_data.text == "V":
                        excel_cell.font = Font(bold=True, size=14)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    val_str = str(cell.value) if cell.value is not None else ""
                    if len(val_str) > max_length:
                        max_length = len(val_str)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Ensure directory exists
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        wb.save(output_path)



def convert_table_to_grid_cells(table: TableStructure) -> List[List]:
    """
    Convert TableStructure to old GridCell format for backward compatibility.
    
    Args:
        table: TableStructure to convert
        
    Returns:
        List of rows, each row is a list of GridCell objects
    """
    from verb_centered_model import GridCell
    
    grid_rows = []
    for row in table.rows:
        if not row:
            grid_rows.append([])
            continue
        
        grid_row = []
        for cell in row:
            grid_cell = GridCell(
                text=cell.text,
                value=cell.value,
                cell_type=cell.cell_type,
                rich_text=cell.rich_segments if cell.rich_segments else []
            )
            grid_row.append(grid_cell)
        
        grid_rows.append(grid_row)
    
    return grid_rows
